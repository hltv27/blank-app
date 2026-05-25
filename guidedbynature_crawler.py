#!/usr/bin/env python3
"""
Crawler v5 — guidedbynature.pt
Sem pandas, sem lxml — compatível com Termux/Android.

O site renderiza listagens em JavaScript (AJAX), por isso o crawler
descobre e usa a API interna em vez de tentar ler HTML estático.

Estratégia:
  1. Descobrir API analisando ficheiros JS + tentar padrões conhecidos
  2. Usar API para paginar todas as categorias e obter TODOS os items
  3. Fallback: BFS + extracção de URLs em <script>
  4. Guardar debug_listing.html se nada funcionar

Dependências:
    pip install beautifulsoup4 openpyxl cloudscraper

Correr:
    python guidedbynature_crawler.py
"""

try:
    import cloudscraper
    session = cloudscraper.create_scraper()
    session.headers.update({
        "User-Agent":      "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                           "(KHTML, like Gecko) Chrome/136.0.0.0 Safari/537.36",
        "Accept":          "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "pt-PT,pt;q=0.9,en;q=0.7",
        "Accept-Encoding": "gzip, deflate, br",
        "Connection":      "keep-alive",
        "Cache-Control":   "max-age=0",
    })
    print("[Info] cloudscraper activo")
except ImportError:
    import requests
    session = requests.Session()
    session.headers.update({
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                      "(KHTML, like Gecko) Chrome/136.0.0.0 Safari/537.36",
    })
    print("[Info] requests simples")

from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from urllib.parse import urljoin, urlparse, urlencode
from collections import deque
import json
import re
import time

BASE_URL = "https://guidedbynature.pt"
OUTPUT   = "guidedbynature_dados.xlsx"
DELAY    = 0.8

_ITEM_RE = re.compile(
    r"/pt/(poi|event|tour)/[^/?#]+/[^/?#]+/\d+/?$"
    r"|/pt/p/[^/?#]+/\d+/?$"
)
_ITEM_RE_LOOSE = re.compile(r'["\']?(/pt/(poi|event|tour|p)/[^\s"\'<>]+/\d+)["\']?')
_SKIP = ["/static/", "/media/", "/admin/", "/en/", "/es/", "/fr/", "/de/",
         ".pdf", ".jpg", ".jpeg", ".png", ".gif", ".svg", ".zip", ".css", ".ico"]


# ── HTTP ─────────────────────────────────────────────────────────────────────

def get_raw(url, timeout=25, accept=None):
    headers = {}
    if accept:
        headers["Accept"] = accept
    for attempt in range(3):
        try:
            r = session.get(url, timeout=timeout, headers=headers)
            if r.status_code == 404:
                return None
            if r.status_code in (429, 503):
                time.sleep(5 * (attempt + 1))
                continue
            r.raise_for_status()
            return r.text
        except Exception as e:
            if attempt < 2:
                time.sleep(2 ** attempt)
            else:
                print(f"  [ERRO] {url}: {e}")
    return None


def get_json(url, params=None):
    try:
        full = url + ("?" + urlencode(params) if params else "")
        r = session.get(full, timeout=25, headers={"Accept": "application/json"})
        if r.status_code == 200:
            ct = r.headers.get("content-type", "")
            if "json" in ct or r.text.strip().startswith(("{", "[")):
                return r.json()
    except Exception:
        pass
    return None


def get_soup(url):
    html = get_raw(url)
    return BeautifulSoup(html, "html.parser") if html else None


def _text(el):
    return el.get_text(" ", strip=True) if el else ""


# ── Normalização de URL ───────────────────────────────────────────────────────

def _normalize(url):
    try:
        p = urlparse(url)
    except Exception:
        return ""
    netloc = (p.netloc or "").replace("www.", "")
    base_n = BASE_URL.split("//")[-1].replace("www.", "")
    if netloc and netloc != base_n:
        return ""
    if p.scheme and p.scheme not in ("http", "https"):
        return ""
    path = p.path or "/"
    if not path.endswith("/"):
        path += "/"
    q = ""
    m = re.search(r"page=(\d+)", p.query or "")
    if m and int(m.group(1)) > 1:
        q = f"?page={m.group(1)}"
    return f"{BASE_URL}{path}{q}"


def _is_item(url):
    return bool(_ITEM_RE.search(url))


def _is_crawlable(url):
    if not url.startswith(BASE_URL):
        return False
    path = urlparse(url).path.lower()
    return not any(s in path for s in _SKIP)


# ══════════════════════════════════════════════════════════════════════════════
# FASE 1 — Descoberta de API
# ══════════════════════════════════════════════════════════════════════════════

# Padrões REST a tentar (em ordem de probabilidade)
_API_ROOTS = [
    "/api/v1/", "/api/v2/", "/api/", "/api/v3/",
    "/rest/v1/", "/rest/",
    "/pt/api/v1/", "/pt/api/",
]

_API_RESOURCES = [
    "poi", "pois", "point-of-interest", "points-of-interest",
    "place", "places", "location", "locations",
    "event", "events", "tour", "tours",
    "item", "items", "listing", "listings",
    "attraction", "attractions",
]

# Parâmetros de paginação a tentar
_PAGE_PARAMS = [
    {"page": 1, "page_size": 100},
    {"page": 1, "per_page": 100},
    {"page": 1, "limit": 100},
    {"offset": 0, "limit": 100},
    {"p": 1, "n": 100},
]


def _probe_api_endpoint(url):
    """Testa um endpoint e retorna (url_base, data) se for uma lista de items."""
    for params in _PAGE_PARAMS:
        data = get_json(url, params)
        if not data:
            continue
        # Aceitar lista directa ou {results: [...], count: N}
        if isinstance(data, list) and len(data) > 0:
            return url, params, data
        if isinstance(data, dict):
            for key in ("results", "items", "data", "objects", "content", "list",
                        "pois", "places", "locations", "events", "tours"):
                if key in data and isinstance(data[key], list) and len(data[key]) > 0:
                    return url, params, data[key]
    return None, None, None


def _find_api_in_js(page_html):
    """Analisa ficheiros JS da página à procura de endpoints de API."""
    soup = BeautifulSoup(page_html, "html.parser")
    candidates = set()

    # Padrões de fetch/axios em JS inline
    for script in soup.find_all("script"):
        text = script.string or ""
        for m in re.finditer(r'(?:fetch|axios\.(?:get|post))\s*\(\s*["\`]([^"\'`\s]+)["\`]', text):
            u = m.group(1)
            if u.startswith("/") or BASE_URL in u:
                candidates.add(u if u.startswith("http") else BASE_URL + u)

    # Ficheiros JS externos (bundle, app, main, chunk)
    js_srcs = [
        urljoin(BASE_URL, s["src"])
        for s in soup.find_all("script", src=True)
        if re.search(r"(main|app|bundle|chunk|vendor|index)", s["src"], re.I)
    ]
    print(f"  [JS] A analisar {len(js_srcs)} ficheiros JS...")

    for js_url in js_srcs[:8]:   # limite para não demorar demasiado
        js = get_raw(js_url, timeout=30)
        if not js:
            continue
        # fetch/axios calls
        for m in re.finditer(r'(?:fetch|axios\.(?:get|post))\s*\(\s*["\`]([^"\'`\s]+)["\`]', js):
            u = m.group(1)
            if "/api/" in u or "/rest/" in u:
                candidates.add(u if u.startswith("http") else BASE_URL + u)
        # strings que parecem API paths
        for m in re.finditer(r'["\`](/(?:api|rest)/[^"\'`\s?#]{3,50})["\`]', js):
            u = BASE_URL + m.group(1)
            candidates.add(u)

    return candidates


def discover_api():
    """Tenta encontrar a API do site. Retorna (endpoint_url, page_params) ou (None, None)."""
    print("[API] A tentar descobrir endpoint de API...")

    # 1. Tentar todas as combinações root + resource
    for root in _API_ROOTS:
        for res in _API_RESOURCES:
            url = BASE_URL + root + res + "/"
            ep, params, data = _probe_api_endpoint(url)
            if ep:
                print(f"  [API] ✓ Encontrado: {ep} → {len(data)} items na primeira página")
                return ep, params
        time.sleep(0.2)

    # 2. Tentar descoberta via raiz da API
    for root in _API_ROOTS:
        data = get_json(BASE_URL + root)
        if isinstance(data, dict):
            print(f"  [API] Raiz {root} devolve JSON: {list(data.keys())[:10]}")
            # Tentar cada chave como sub-endpoint
            for key, val in data.items():
                if isinstance(val, str) and val.startswith("http"):
                    ep, params, items = _probe_api_endpoint(val)
                    if ep:
                        print(f"  [API] ✓ {ep}")
                        return ep, params

    # 3. Analisar JS da homepage
    print("  [API] A analisar ficheiros JS da homepage...")
    html = get_raw(BASE_URL + "/pt/")
    if html:
        candidates = _find_api_in_js(html)
        print(f"  [API] {len(candidates)} candidatos em JS")
        for url in sorted(candidates):
            ep, params, data = _probe_api_endpoint(url)
            if ep:
                print(f"  [API] ✓ Encontrado via JS: {ep}")
                return ep, params

    print("  [API] Nenhuma API encontrada — a usar BFS")
    return None, None


# ── Paginar API para obter TODOS os items ─────────────────────────────────────

def _extract_url_from_api_item(item):
    """Extrai a URL ou ID de um item retornado pela API."""
    if isinstance(item, dict):
        for key in ("url", "link", "href", "permalink", "absolute_url", "get_absolute_url"):
            if key in item and isinstance(item[key], str):
                return item[key]
        for key in ("id", "pk", "slug"):
            if key in item:
                return str(item[key])
    return str(item)


def fetch_all_via_api(endpoint, page_params):
    """Pagina completamente um endpoint de API e retorna todos os items."""
    all_items = []
    params = dict(page_params)

    # Determinar o campo de página
    page_key   = next((k for k in params if k in ("page", "p")), None)
    offset_key = next((k for k in params if k == "offset"), None)
    limit_key  = next((k for k in params if k in ("limit", "per_page", "page_size", "n")), None)
    limit_val  = params.get(limit_key, 100) if limit_key else 100

    page = 1
    while True:
        data = get_json(endpoint, params)
        if not data:
            break

        items = []
        if isinstance(data, list):
            items = data
        elif isinstance(data, dict):
            for key in ("results", "items", "data", "objects", "content", "list",
                        "pois", "places", "locations", "events", "tours"):
                if key in data and isinstance(data[key], list):
                    items = data[key]
                    break

        if not items:
            break

        all_items.extend(items)
        print(f"  [API pág {page}] +{len(items)} items  (total: {len(all_items)})")

        # Verificar se há mais
        if len(items) < limit_val:
            break
        if isinstance(data, dict) and "next" in data:
            if not data["next"]:
                break

        # Avançar para a próxima página
        page += 1
        if page_key:
            params[page_key] = page
        elif offset_key and limit_key:
            params[offset_key] += limit_val
        else:
            break
        time.sleep(DELAY)

    print(f"  [API] Total: {len(all_items)} items")
    return all_items


# ══════════════════════════════════════════════════════════════════════════════
# FASE 2 — Sitemap XML
# ══════════════════════════════════════════════════════════════════════════════

def _parse_sitemap(url, item_urls, visited, depth=0):
    if url in visited or depth > 6:
        return
    visited.add(url)
    html = get_raw(url, timeout=20)
    if not html:
        return
    locs = re.findall(r"<loc>\s*([^<]+)\s*</loc>", html)
    if "<sitemapindex" in html or (any(l.endswith(".xml") for l in locs)):
        sub = [l.strip() for l in locs if l.strip().endswith(".xml")]
        print(f"  [Sitemap] índice com {len(sub)} sub-sitemaps")
        for s in sub:
            _parse_sitemap(s, item_urls, visited, depth + 1)
        return
    before = len(item_urls)
    for loc in locs:
        loc = loc.strip()
        if _ITEM_RE.search(loc):
            item_urls.add(loc.rstrip("/") + "/")
    added = len(item_urls) - before
    if added:
        print(f"  [Sitemap] +{added} ({len(item_urls)} total) — {url}")
    time.sleep(0.3)


def discover_via_sitemap():
    item_urls, visited = set(), set()
    html = get_raw(BASE_URL + "/robots.txt", timeout=10)
    if html:
        for m in re.finditer(r"(?i)Sitemap:\s*(\S+)", html):
            _parse_sitemap(m.group(1).strip(), item_urls, visited)
    if not item_urls:
        for path in ["/sitemap.xml", "/sitemap_index.xml", "/pt/sitemap.xml"]:
            _parse_sitemap(BASE_URL + path, item_urls, visited)
            if item_urls:
                break
    return item_urls


# ══════════════════════════════════════════════════════════════════════════════
# FASE 3 — BFS com extracção de URLs em scripts
# ══════════════════════════════════════════════════════════════════════════════

def _urls_from_html(html, base):
    """Extrai URLs de items de qualquer lugar no HTML: links, scripts, data-attrs."""
    found = set()
    for m in _ITEM_RE_LOOSE.finditer(html):
        raw  = m.group(1)
        norm = _normalize(urljoin(base, raw))
        if norm and _is_item(norm):
            found.add(norm)
    return found


def discover_via_bfs(known_items):
    seeds = [BASE_URL + "/pt/", BASE_URL + "/",
             BASE_URL + "/pt/poi/", BASE_URL + "/pt/event/",
             BASE_URL + "/pt/tour/", BASE_URL + "/pt/p/"]
    queue     = deque(seeds)
    visited   = set()
    item_urls = set(known_items)
    nav_count = 0

    print(f"[BFS] {len(seeds)} sementes | {len(known_items)} items já conhecidos")

    while queue and nav_count < 8000:
        url  = queue.popleft()
        norm = _normalize(url)
        if not norm or norm in visited or not _is_crawlable(norm):
            continue
        visited.add(norm)
        nav_count += 1

        if _is_item(norm):
            item_urls.add(norm)
            continue

        if nav_count % 20 == 0 or nav_count <= 5:
            print(f"  [BFS {nav_count:4}] {norm}  |  items: {len(item_urls)}")

        html = get_raw(norm)
        if not html:
            time.sleep(DELAY)
            continue

        # Salvar primeira página de listagem para debug
        if nav_count == 2:
            with open("debug_listing.html", "w", encoding="utf-8") as f:
                f.write(html)
            print("  [Debug] debug_listing.html guardado")

        # Extrair items de qualquer sítio no HTML
        for u in _urls_from_html(html, norm):
            item_urls.add(u)

        # Seguir links internos (navegação + paginação)
        soup = BeautifulSoup(html, "html.parser")
        for a in soup.find_all("a", href=True):
            href = a["href"].strip()
            if not href or href.startswith(("javascript", "mailto", "tel")):
                continue
            full  = urljoin(norm, href).split("#")[0]
            norm2 = _normalize(full)
            if not norm2 or not _is_crawlable(norm2) or norm2 in visited:
                continue
            if _is_item(norm2):
                item_urls.add(norm2)
            else:
                queue.append(norm2)

        time.sleep(DELAY)

    print(f"[BFS] {nav_count} páginas | {len(item_urls)} items")
    return item_urls


# ══════════════════════════════════════════════════════════════════════════════
# Extracção de dados de cada item
# ══════════════════════════════════════════════════════════════════════════════

def _extract_from_api_item(item: dict, base_url_guess="") -> dict:
    """Converte um item da API directamente para linha Excel."""
    row = {}
    url = _extract_url_from_api_item(item)
    row["url"] = urljoin(BASE_URL, url) if not url.startswith("http") else url

    # Tentar mapear campos comuns da API
    for dest, srcs in {
        "id_poi":    ("id", "pk", "uid"),
        "nome":      ("name", "title", "nome", "label"),
        "descricao": ("description", "body", "descricao", "summary", "excerpt"),
        "tipo_url":  ("type", "content_type", "kind"),
        "categoria": ("category", "categories", "categoria", "type_label"),
        "slug":      ("slug",),
        "latitude":  ("lat", "latitude"),
        "longitude": ("lng", "lon", "longitude"),
        "morada":    ("address", "street", "morada"),
        "localidade":("city", "locality", "localidade", "municipality"),
        "regiao":    ("region", "district", "regiao"),
        "telefone":  ("phone", "telephone", "tel"),
        "email":     ("email",),
        "website":   ("website", "url", "homepage"),
        "imagem":    ("image", "photo", "thumbnail", "cover"),
    }.items():
        for src in srcs:
            if src in item and item[src]:
                val = item[src]
                if isinstance(val, (dict, list)):
                    val = json.dumps(val, ensure_ascii=False)
                row[dest] = str(val)
                break

    # Campos extras: tudo o resto
    for k, v in item.items():
        if k not in row:
            row[f"api_{k}"] = str(v) if not isinstance(v, (dict, list)) else json.dumps(v, ensure_ascii=False)[:500]

    return row


def extract_jsonld(soup) -> dict:
    for script in soup.find_all("script", type="application/ld+json"):
        try:
            data = json.loads(script.string or "")
            if isinstance(data, list):
                data = data[0]
            if isinstance(data, dict):
                return data
        except Exception:
            pass
    return {}


def extract_meta(soup) -> dict:
    meta = {}
    for tag in soup.find_all("meta"):
        key = tag.get("property") or tag.get("name") or ""
        val = tag.get("content", "").strip()
        if key and val:
            meta[key] = val
    return meta


def extract_item(url: str) -> dict | None:
    soup = get_soup(url)
    if not soup:
        return None

    row: dict = {"url": url}
    parts = url.rstrip("/").split("/")
    row["id_poi"] = parts[-1] if parts[-1].isdigit() else ""

    if "/pt/poi/" in url:
        row["tipo_url"]  = "poi"
        row["slug"]      = parts[-2] if len(parts) >= 2 else ""
        row["categoria"] = parts[-3] if len(parts) >= 3 else ""
    elif "/pt/event/" in url:
        row["tipo_url"]  = "evento"
        row["slug"]      = parts[-2] if len(parts) >= 2 else ""
        row["categoria"] = parts[-3] if len(parts) >= 3 else ""
    elif "/pt/tour/" in url:
        row["tipo_url"]  = "tour"
        row["slug"]      = parts[-2] if len(parts) >= 2 else ""
        row["categoria"] = parts[-3] if len(parts) >= 3 else ""
    elif "/pt/p/" in url:
        row["tipo_url"]  = parts[-2] if len(parts) >= 2 else "p"
        row["slug"]      = ""
        row["categoria"] = row["tipo_url"]

    jld = extract_jsonld(soup)
    if jld:
        row["nome"]      = jld.get("name", "")
        row["descricao"] = jld.get("description", "")
        row["tipo"]      = jld.get("@type", "")
        row["website"]   = jld.get("url", "")
        row["telefone"]  = jld.get("telephone", "")
        row["email"]     = jld.get("email", "")
        addr = jld.get("address", {})
        if isinstance(addr, dict):
            row["morada"]     = addr.get("streetAddress", "")
            row["localidade"] = addr.get("addressLocality", "")
            row["regiao"]     = addr.get("addressRegion", "")
            row["pais"]       = addr.get("addressCountry", "")
        elif isinstance(addr, str):
            row["morada"] = addr
        geo = jld.get("geo", {})
        if isinstance(geo, dict):
            row["latitude"]  = geo.get("latitude", "")
            row["longitude"] = geo.get("longitude", "")
        row["preco"]         = str(jld.get("priceRange", "") or jld.get("price", ""))
        row["horario"]       = str(jld.get("openingHours", "") or jld.get("openingHoursSpecification", ""))
        row["imagem_jsonld"] = str(jld.get("image", ""))

    meta = extract_meta(soup)
    if not row.get("nome"):
        row["nome"] = meta.get("og:title", meta.get("title", ""))
    if not row.get("descricao"):
        row["descricao"] = meta.get("og:description", meta.get("description", ""))
    row["og_image"] = meta.get("og:image", "")

    h1 = soup.find("h1")
    if h1:
        row["h1"] = _text(h1)

    pares = []
    for dl in soup.find_all("dl"):
        for dt, dd in zip(dl.find_all("dt"), dl.find_all("dd")):
            pares.append((_text(dt), _text(dd)))
    for tr in soup.find_all("tr"):
        for th, td in zip(tr.find_all("th"), tr.find_all("td")):
            pares.append((_text(th), _text(td)))
    for el in soup.find_all(class_=re.compile(r"field|detail|info|meta|attr|prop", re.I)):
        lbl = el.find(["label","strong","b","dt","th","span"],
                      class_=re.compile(r"label|key|title|name", re.I))
        val = el.find(["span","p","div","dd","td"],
                      class_=re.compile(r"value|content|data|text", re.I))
        if lbl and val:
            pares.append((_text(lbl), _text(val)))
    for lbl, val in pares:
        if not lbl or not val or lbl == val:
            continue
        key = "campo_" + re.sub(r"[^\w]", "_", lbl.lower().strip("_"))[:40]
        if key not in row:
            row[key] = val

    imgs = []
    for img in soup.find_all("img", src=True):
        src = img["src"]
        if any(x in src for x in ["/photo", "/image", "/media", "/upload", "/poi", "/tour", "/event"]):
            imgs.append(urljoin(BASE_URL, src))
    row["imagens"] = " | ".join(dict.fromkeys(imgs))

    if not row.get("descricao"):
        for tag in soup.find_all(["p","div"], class_=re.compile(r"desc|content|body|text|about", re.I)):
            txt = _text(tag)
            if len(txt) > 120:
                row["descricao"] = txt[:3000]
                break

    tags = []
    for tag in soup.find_all(class_=re.compile(r"\btag\b|\bbadge\b|\blabel\b|\bchip\b", re.I)):
        t = _text(tag)
        if t and len(t) < 60:
            tags.append(t)
    row["tags"] = " | ".join(dict.fromkeys(tags))

    sociais = []
    for a in soup.find_all("a", href=True):
        href = a["href"]
        if any(s in href for s in ["facebook.com","instagram.com","twitter.com",
                                    "youtube.com","tiktok.com","linkedin.com"]):
            sociais.append(href)
    row["redes_sociais"] = " | ".join(dict.fromkeys(sociais))

    return row


# ══════════════════════════════════════════════════════════════════════════════
# Excel
# ══════════════════════════════════════════════════════════════════════════════

_COLS_PRIORITY = [
    "id_poi", "nome", "h1", "tipo_url", "categoria", "slug", "tipo",
    "descricao", "morada", "localidade", "regiao", "pais",
    "latitude", "longitude", "telefone", "email", "website",
    "preco", "horario", "imagem_jsonld", "og_image", "imagens",
    "tags", "redes_sociais", "url",
]

_HDR_FILL = PatternFill("solid", fgColor="1F4E79")
_HDR_FONT = Font(color="FFFFFF", bold=True)
_X_FILL   = PatternFill("solid", fgColor="E2EFDA")


def _write_sheet(ws, headers, rows, presence=False):
    ws.append(headers)
    for i in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=i)
        c.fill = _HDR_FILL; c.font = _HDR_FONT
        c.alignment = Alignment(horizontal="center")
    for row in rows:
        vals = []
        for col in headers:
            v = str(row.get(col, "") or "").strip()
            vals.append("X" if (presence and v) else ("" if presence else v))
        ws.append(vals)
    if presence:
        for r in ws.iter_rows(min_row=2):
            for cell in r:
                if cell.value == "X":
                    cell.fill = _X_FILL
                    cell.alignment = Alignment(horizontal="center")
    for i, name in enumerate(headers, 1):
        mx = len(name)
        for r in ws.iter_rows(min_row=2, min_col=i, max_col=i):
            for cell in r:
                if cell.value:
                    mx = max(mx, min(len(str(cell.value)), 60))
        ws.column_dimensions[get_column_letter(i)].width = mx + 2


def _write_resumo(ws, dados):
    from collections import Counter
    ws.append(["tipo_url", "categoria", "total"])
    for i in range(1, 4):
        c = ws.cell(row=1, column=i)
        c.fill = _HDR_FILL; c.font = _HDR_FONT
        c.alignment = Alignment(horizontal="center")
    counter = Counter((str(r.get("tipo_url","")), str(r.get("categoria",""))) for r in dados)
    for (t, c), n in sorted(counter.items()):
        ws.append([t, c, n])
    for i in range(1, 4):
        ws.column_dimensions[get_column_letter(i)].width = 25


def save_excel(dados, output=OUTPUT):
    if not dados:
        print("[Erro] Sem dados."); return
    all_keys, seen = [], set()
    for row in dados:
        for k in row:
            if k not in seen:
                seen.add(k); all_keys.append(k)
    priority = [c for c in _COLS_PRIORITY if c in seen]
    headers  = priority + [c for c in all_keys if c not in priority]

    wb = Workbook()
    ws = wb.active; ws.title = "Dados"
    _write_sheet(ws, headers, dados)
    ws2 = wb.create_sheet("Presença")
    _write_sheet(ws2, headers, dados, presence=True)
    ws3 = wb.create_sheet("Resumo")
    _write_resumo(ws3, dados)
    wb.save(output)

    print(f"\n✅  {output}")
    print(f"    {len(dados)} linhas × {len(headers)} colunas")
    print(f"    Sheets: Dados | Presença (X/vazio) | Resumo")


# ══════════════════════════════════════════════════════════════════════════════
# Main
# ══════════════════════════════════════════════════════════════════════════════

def main():
    print("=" * 60)
    print("  guidedbynature.pt — Crawler v5")
    print("=" * 60)

    try:
        r = session.get(BASE_URL + "/pt/", timeout=20)
        print(f"[Session] HTTP {r.status_code}")
    except Exception as e:
        print(f"[Session] {e}")

    dados = []

    # ── Fase 1: API ──────────────────────────────────────────────────────────
    print("\n[Fase 1] Descoberta de API...")
    api_endpoint, api_params = discover_api()

    if api_endpoint:
        print(f"\n[Fase 1b] A paginar API: {api_endpoint}")
        api_items = fetch_all_via_api(api_endpoint, api_params)
        if api_items:
            print(f"[API] A converter {len(api_items)} items para Excel...")
            for item in api_items:
                row = _extract_from_api_item(item)
                if row:
                    dados.append(row)

    # ── Fase 2: Sitemap ──────────────────────────────────────────────────────
    print("\n[Fase 2] Sitemap XML...")
    sitemap_urls = discover_via_sitemap()
    print(f"  → {len(sitemap_urls)} items via sitemap")

    # ── Fase 3: BFS ──────────────────────────────────────────────────────────
    print("\n[Fase 3] BFS + extracção de scripts...")
    all_item_urls = discover_via_bfs(sitemap_urls)
    print(f"  → {len(all_item_urls)} URLs de items no total")

    if not all_item_urls and not dados:
        print("\n[Aviso] Nenhum item encontrado.")
        print("Verifica debug_listing.html para perceber o que o servidor devolve.")
        print("Pode ser necessário Selenium para sites 100% JS.")
        return

    # ── Fase 4: Extracção HTML item a item ───────────────────────────────────
    urls_already = {r.get("url","") for r in dados}
    urls_to_fetch = [u for u in sorted(all_item_urls) if u not in urls_already]

    if urls_to_fetch:
        total = len(urls_to_fetch)
        print(f"\n[Fase 4] Extracção HTML: {total} items\n")
        for i, url in enumerate(urls_to_fetch, 1):
            if i <= 10 or i % 50 == 0:
                print(f"[{i:5}/{total}] {url}")
            row = extract_item(url)
            if row:
                dados.append(row)
            if i % 300 == 0 and dados:
                save_excel(dados, OUTPUT.replace(".xlsx", f"_parcial_{i}.xlsx"))
            time.sleep(DELAY)

    save_excel(dados)


if __name__ == "__main__":
    main()
