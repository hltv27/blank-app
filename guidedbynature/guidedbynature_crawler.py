#!/usr/bin/env python3
"""
Crawler v6 — guidedbynature.pt  (Playwright + openpyxl, sem pandas)
Usa browser real (Chromium headless) para executar JavaScript.
Encontra TODOS os items — hotéis, restaurantes, POIs, tours, eventos.

Instalar dependências:
    pip install playwright openpyxl
    playwright install chromium

Se playwright install chromium falhar no Termux:
    pkg install chromium
    pip install playwright
    (o script detecta automaticamente o chromium do sistema)

Correr:
    python guidedbynature_crawler.py

Ficheiro gerado: guidedbynature_dados.xlsx  (Dados | Presença | Resumo)
"""

import asyncio
import json
import re
from collections import Counter
from urllib.parse import urljoin

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from playwright.async_api import async_playwright, Page, BrowserContext

BASE_URL = "https://guidedbynature.pt"
OUTPUT   = "guidedbynature_dados.xlsx"
DELAY    = 0.6   # segundos entre páginas de detalhe

# Todas as categorias conhecidas do site
CATEGORIAS = [
    # PASSEAR
    ("tour", "caminhadas"),
    ("tour", "bicicleta"),
    ("tour", "roteiros-tematicos"),
    ("tour", "caminhada-de-longa-distancia"),
    ("tour", "caminhada"),
    ("tour", "btt"),
    ("tour", "ciclismo"),
    ("tour", "trail"),
    ("tour", "canoagem"),
    ("tour", "escalada"),
    ("tour", "equitacao"),
    ("tour", "outros"),
    # MERGULHAR
    ("poi", "praias-ou-piscinas"),
    ("poi", "docas-cais-e-fluvinas"),
    ("poi", "termas-e-spa"),
    # CONHECER
    ("poi", "paisagens"),
    ("poi", "historia-e-cultura"),
    ("poi", "gastronomia"),
    ("poi", "miradouros"),
    ("poi", "monumentos"),
    ("poi", "museus"),
    ("poi", "aldeias"),
    ("poi", "cascatas"),
    ("poi", "lagos"),
    ("poi", "praias-fluviais"),
    ("poi", "outras-paisagens"),
    ("poi", "pontos-de-interesse"),
    ("poi", "parques"),
    ("poi", "ponte"),
    ("poi", "centro-cultural"),
    ("poi", "igreja"),
    # PLANEAR
    ("poi", "onde-comer"),
    ("poi", "onde-dormir"),
    ("poi", "postos-de-turismo"),
    # EVENTOS
    ("event", "feiras-e-tradicoes"),
    ("event", "festivais"),
    ("event", "mercados"),
    ("event", "concertos"),
    ("event", "desporto"),
    ("event", "cultura"),
    ("event", "gastronomia"),
    ("event", "natureza"),
    ("event", "outros"),
    ("event", "corrida"),
    # PARTICIPAR
    ("p", "desafios"),
    ("p", "experiencias-guiadas"),
    ("p", "comprar-produtos-locais"),
    ("p", "eventos"),
    ("p", "noticias"),
    ("p", "percursos"),
    ("p", "alojamento"),
    ("p", "restauracao"),
    ("p", "servicos"),
    ("p", "outros"),
    ("p", "contactos"),
    ("p", "praias-fluviais-ou-piscinas"),
]

_ITEM_RE = re.compile(r'href="(/pt/(poi|event|tour|p)/[^"]+/\d+/)"')


# ── Descoberta de URLs (com browser real) ─────────────────────────────────────

async def discover_urls(context: BrowserContext) -> list[str]:
    """
    Percorre cada categoria e pagina até ao fim.
    O browser executa o JavaScript, por isso os cards são visíveis.
    """
    urls: set[str] = set()
    page = await context.new_page()

    # Também tenta descobrir categorias automaticamente na navegação
    try:
        await page.goto(BASE_URL + "/pt/", wait_until="domcontentloaded", timeout=30000)
        await page.wait_for_timeout(1500)
        content = await page.content()
        nav_cats = set()
        for m in re.finditer(r'href="(/pt/(poi|event|tour|p)/([^/"]+))"', content):
            tipo, slug = m.group(2), m.group(3)
            if not slug.isdigit():
                nav_cats.add((tipo, slug))
        if nav_cats:
            print(f"[Nav] {len(nav_cats)} categorias encontradas na navegação")
    except Exception:
        nav_cats = set()

    # Combina categorias da nav com as hardcoded
    all_cats = list(dict.fromkeys(list(nav_cats) + CATEGORIAS))

    for tipo, cat in all_cats:
        cat_url = f"{BASE_URL}/pt/{tipo}/{cat}/"
        cat_found = 0

        for pg in range(1, 200):   # até 200 páginas por categoria
            url = cat_url if pg == 1 else f"{cat_url}?page={pg}"
            try:
                resp = await page.goto(url, wait_until="domcontentloaded", timeout=30000)
                if not resp or resp.status == 404:
                    break
            except Exception as e:
                print(f"  [ERRO nav] {url}: {e}")
                break

            await page.wait_for_timeout(1200)
            content = await page.content()

            before = len(urls)
            for m in _ITEM_RE.finditer(content):
                full = urljoin(BASE_URL, m.group(1)).rstrip("/") + "/"
                urls.add(full)
            new = len(urls) - before
            cat_found += new

            if new == 0:   # sem novos links → última página
                break

        if cat_found:
            print(f"  [{tipo}/{cat}] {cat_found} items")

    await page.close()
    print(f"\n[Descoberta] {len(urls)} URLs no total\n")
    return list(urls)


# ── Extracção de dados de cada página ─────────────────────────────────────────

async def extract_page(page: Page, url: str) -> dict | None:
    try:
        resp = await page.goto(url, wait_until="domcontentloaded", timeout=30000)
        if not resp or resp.status >= 400:
            return None
    except Exception as e:
        print(f"  [ERRO] {url}: {e}")
        return None

    await page.wait_for_timeout(800)

    row: dict = {"url": url}
    parts = url.rstrip("/").split("/")
    row["id_poi"] = parts[-1] if parts[-1].isdigit() else ""

    if "/pt/poi/" in url:
        row["tipo_url"]  = "poi"
        row["slug"]      = parts[-2] if len(parts) >= 2 else ""
        row["categoria"] = parts[-3] if len(parts) >= 3 else ""
    elif "/pt/event/" in url:
        row["tipo_url"]  = "evento"
        row["slug"]      = parts[-2] if len(parts) >= 2 else ""
        row["categoria"] = parts[-3] if len(parts) >= 3 else ""
    elif "/pt/tour/" in url:
        row["tipo_url"]  = "tour"
        row["slug"]      = parts[-2] if len(parts) >= 2 else ""
        row["categoria"] = parts[-3] if len(parts) >= 3 else ""
    elif "/pt/p/" in url:
        row["tipo_url"]  = parts[-2] if len(parts) >= 2 else "p"
        row["slug"]      = ""
        row["categoria"] = row["tipo_url"]

    # JSON-LD (dados estruturados — mais fiáveis)
    for tag in await page.query_selector_all('script[type="application/ld+json"]'):
        try:
            data = json.loads(await tag.inner_text())
            if isinstance(data, list):
                data = data[0]
            if not isinstance(data, dict):
                continue
            row["nome"]      = data.get("name", "")
            row["descricao"] = data.get("description", "")
            row["tipo"]      = data.get("@type", "")
            row["website"]   = data.get("url", "")
            row["telefone"]  = data.get("telephone", "")
            row["email"]     = data.get("email", "")
            addr = data.get("address", {})
            if isinstance(addr, dict):
                row["morada"]     = addr.get("streetAddress", "")
                row["localidade"] = addr.get("addressLocality", "")
                row["regiao"]     = addr.get("addressRegion", "")
                row["pais"]       = addr.get("addressCountry", "")
            elif isinstance(addr, str):
                row["morada"] = addr
            geo = data.get("geo", {})
            if isinstance(geo, dict):
                row["latitude"]  = geo.get("latitude", "")
                row["longitude"] = geo.get("longitude", "")
            row["preco"]         = str(data.get("priceRange", "") or data.get("price", ""))
            row["horario"]       = str(data.get("openingHours", "") or "")
            row["imagem_jsonld"] = str(data.get("image", ""))
            break
        except Exception:
            pass

    # Open Graph como fallback
    if not row.get("nome"):
        el = await page.query_selector('meta[property="og:title"]')
        if el:
            row["nome"] = (await el.get_attribute("content") or "").strip()
    if not row.get("descricao"):
        el = await page.query_selector('meta[property="og:description"], meta[name="description"]')
        if el:
            row["descricao"] = (await el.get_attribute("content") or "").strip()
    og_img = await page.query_selector('meta[property="og:image"]')
    if og_img:
        row["og_image"] = (await og_img.get_attribute("content") or "").strip()

    # H1
    h1 = await page.query_selector("h1")
    if h1:
        row["h1"] = (await h1.inner_text()).strip()
        if not row.get("nome"):
            row["nome"] = row["h1"]

    # Pares label → valor (dl/dt/dd, tabelas)
    content = await page.content()
    from bs4 import BeautifulSoup
    soup = BeautifulSoup(content, "html.parser")

    pares = []
    for dl in soup.find_all("dl"):
        for dt, dd in zip(dl.find_all("dt"), dl.find_all("dd")):
            t_dt = dt.get_text(" ", strip=True)
            t_dd = dd.get_text(" ", strip=True)
            if t_dt and t_dd:
                pares.append((t_dt, t_dd))
    for tr in soup.find_all("tr"):
        for th, td in zip(tr.find_all("th"), tr.find_all("td")):
            t_th = th.get_text(" ", strip=True)
            t_td = td.get_text(" ", strip=True)
            if t_th and t_td:
                pares.append((t_th, t_td))

    for lbl, val in pares:
        if lbl == val:
            continue
        key = "campo_" + re.sub(r"[^\w]", "_", lbl.lower().strip("_"))[:40]
        if key not in row:
            row[key] = val

    # Imagens
    imgs = []
    for img in soup.find_all("img", src=True):
        src = img["src"]
        if any(x in src for x in ["/photo", "/image", "/media", "/upload", "/poi", "/tour", "/event"]):
            imgs.append(urljoin(BASE_URL, src))
    row["imagens"] = " | ".join(dict.fromkeys(imgs))

    # Tags
    tags = []
    for el in soup.find_all(class_=re.compile(r"\btag\b|\bbadge\b|\bchip\b", re.I)):
        t = el.get_text(" ", strip=True)
        if t and len(t) < 60:
            tags.append(t)
    row["tags"] = " | ".join(dict.fromkeys(tags))

    # Redes sociais
    sociais = []
    for a in soup.find_all("a", href=True):
        href = a["href"]
        if any(s in href for s in ["facebook.com", "instagram.com", "twitter.com",
                                    "youtube.com", "tiktok.com", "linkedin.com"]):
            sociais.append(href)
    row["redes_sociais"] = " | ".join(dict.fromkeys(sociais))

    return row


# ── Excel (openpyxl directo, sem pandas) ─────────────────────────────────────

_COLS_PRIORITY = [
    "id_poi", "nome", "h1", "tipo_url", "categoria", "slug", "tipo",
    "descricao", "morada", "localidade", "regiao", "pais",
    "latitude", "longitude", "telefone", "email", "website",
    "preco", "horario", "imagem_jsonld", "og_image", "imagens",
    "tags", "redes_sociais", "url",
]

_HDR_FILL = PatternFill("solid", fgColor="1F4E79")
_HDR_FONT = Font(color="FFFFFF", bold=True)
_X_FILL   = PatternFill("solid", fgColor="E2EFDA")


def _write_sheet(ws, headers, rows, presence=False):
    ws.append(headers)
    for i in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=i)
        c.fill = _HDR_FILL
        c.font = _HDR_FONT
        c.alignment = Alignment(horizontal="center")
    for row in rows:
        vals = []
        for col in headers:
            v = str(row.get(col, "") or "").strip()
            vals.append("X" if (presence and v) else ("" if presence else v))
        ws.append(vals)
    if presence:
        for r in ws.iter_rows(min_row=2):
            for cell in r:
                if cell.value == "X":
                    cell.fill = _X_FILL
                    cell.alignment = Alignment(horizontal="center")
    for i, name in enumerate(headers, 1):
        mx = len(name)
        for r in ws.iter_rows(min_row=2, min_col=i, max_col=i):
            for cell in r:
                if cell.value:
                    mx = max(mx, min(len(str(cell.value)), 60))
        ws.column_dimensions[get_column_letter(i)].width = mx + 2


def save_excel(dados, output=OUTPUT):
    if not dados:
        print("[Erro] Sem dados para guardar.")
        return

    all_keys, seen = [], set()
    for row in dados:
        for k in row:
            if k not in seen:
                seen.add(k)
                all_keys.append(k)

    priority = [c for c in _COLS_PRIORITY if c in seen]
    headers  = priority + [c for c in all_keys if c not in priority]

    wb = Workbook()

    ws = wb.active
    ws.title = "Dados"
    _write_sheet(ws, headers, dados)

    ws2 = wb.create_sheet("Presença")
    _write_sheet(ws2, headers, dados, presence=True)

    ws3 = wb.create_sheet("Resumo")
    ws3.append(["tipo_url", "categoria", "total"])
    for i in range(1, 4):
        c = ws3.cell(row=1, column=i)
        c.fill = _HDR_FILL; c.font = _HDR_FONT
        c.alignment = Alignment(horizontal="center")
    counter = Counter(
        (str(r.get("tipo_url", "")), str(r.get("categoria", ""))) for r in dados
    )
    for (t, c), n in sorted(counter.items()):
        ws3.append([t, c, n])
    for i in range(1, 4):
        ws3.column_dimensions[get_column_letter(i)].width = 25

    wb.save(output)
    print(f"\n✅  {output}")
    print(f"    {len(dados)} linhas × {len(headers)} colunas")
    print(f"    Sheets: Dados | Presença (X/vazio) | Resumo")


# ── Main ──────────────────────────────────────────────────────────────────────

async def main():
    print("=" * 60)
    print("  guidedbynature.pt — Crawler v6 (Playwright)")
    print("=" * 60)

    async with async_playwright() as p:
        # Tenta usar chromium do playwright; se não existir, usa o do sistema
        try:
            browser = await p.chromium.launch(headless=True)
        except Exception:
            try:
                browser = await p.chromium.launch(
                    headless=True,
                    executable_path="/usr/bin/chromium-browser"  # Termux pkg install chromium
                )
            except Exception:
                browser = await p.chromium.launch(
                    headless=True,
                    executable_path="/usr/bin/chromium"
                )

        context = await browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                       "AppleWebKit/537.36 (KHTML, like Gecko) "
                       "Chrome/136.0.0.0 Safari/537.36",
            locale="pt-PT",
        )

        # Warm-up (cookies)
        page = await context.new_page()
        try:
            await page.goto(BASE_URL + "/pt/", wait_until="domcontentloaded", timeout=30000)
            await page.wait_for_timeout(1500)
            print("[Session] cookies obtidos\n")
        except Exception as e:
            print(f"[Session] {e}")
        await page.close()

        # Fase 1: descoberta de todos os URLs
        print("[Fase 1] A descobrir todos os items por categoria...\n")
        urls = await discover_urls(context)

        if not urls:
            print("[Aviso] Nenhuma URL descoberta.")
            await browser.close()
            return

        # Fase 2: extracção item a item
        total = len(urls)
        print(f"[Fase 2] A extrair dados de {total} items...\n")
        dados = []
        page = await context.new_page()

        for i, url in enumerate(sorted(urls), 1):
            if i <= 10 or i % 50 == 0:
                print(f"[{i:5}/{total}] {url.rstrip('/').split('/')[-2]}")
            row = await extract_page(page, url)
            if row:
                dados.append(row)

            # Guarda parcialmente a cada 300 items
            if i % 300 == 0 and dados:
                save_excel(dados, OUTPUT.replace(".xlsx", f"_parcial_{i}.xlsx"))

            import asyncio as _a
            await _a.sleep(DELAY)

        await page.close()
        await browser.close()

    save_excel(dados)


if __name__ == "__main__":
    asyncio.run(main())
