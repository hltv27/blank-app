#!/usr/bin/env python3
"""
Descarrega todas as fotos a partir do Excel já gerado.
Não precisa de re-crawl — lê o douroetamega_dados.xlsx existente.

Instalar:
    pip install openpyxl cloudscraper

Correr:
    python download_fotos.py
    python download_fotos.py douroetamega_dados.xlsx   (outro ficheiro)
"""

import os
import re
import sys
import time

try:
    import cloudscraper
    sess = cloudscraper.create_scraper()
except ImportError:
    import requests
    sess = requests.Session()

sess.headers.update({
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                  "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/136.0.0.0 Safari/537.36",
})

from openpyxl import load_workbook

FOTOS_DIR  = "douroetamega_fotos"
DELAY      = 0.3
EXCEL_FILE = sys.argv[1] if len(sys.argv) > 1 else "douroetamega_dados.xlsx"


def _slug_safe(s: str) -> str:
    return re.sub(r"[^\w-]", "_", s or "geral")[:60].strip("_") or "geral"


def main():
    if not os.path.exists(EXCEL_FILE):
        print(f"[Erro] Ficheiro não encontrado: {EXCEL_FILE}")
        return

    print(f"[Fotos] A ler {EXCEL_FILE}…")
    wb = load_workbook(EXCEL_FILE, read_only=True)
    ws = wb.active

    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {h: i for i, h in enumerate(headers) if h}

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()
    print(f"  {len(rows)} linhas carregadas\n")

    def get(field):
        return col.get(field)

    total_dl = 0
    total_skip = 0
    errors = 0

    for i, row in enumerate(rows, 1):
        def val(field):
            idx = get(field)
            return str(row[idx] or "") if idx is not None else ""

        imgs_raw = val("imagens")
        if not imgs_raw:
            continue

        secao  = _slug_safe(val("secao"))
        cat    = _slug_safe(val("categoria"))
        slug   = _slug_safe(val("slug") or val("id") or val("nome") or f"item_{i}")
        folder = os.path.join(FOTOS_DIR, secao, cat, slug)
        os.makedirs(folder, exist_ok=True)

        for img_url in imgs_raw.split(" | "):
            img_url = img_url.strip()
            if not img_url:
                continue
            fname = img_url.split("/")[-1].split("?")[0] or "foto.jpg"
            if "." not in fname[-6:]:
                fname += ".jpg"
            fpath = os.path.join(folder, fname)

            if os.path.exists(fpath):
                total_skip += 1
                continue

            try:
                r = sess.get(img_url, timeout=20)
                if r.ok:
                    with open(fpath, "wb") as f:
                        f.write(r.content)
                    total_dl += 1
                else:
                    errors += 1
            except Exception as e:
                errors += 1

            time.sleep(DELAY)

        if i % 100 == 0:
            print(f"  [{i}/{len(rows)}] {total_dl} fotos | {total_skip} já existiam | {errors} erros")

    print(f"\n✅  Concluído!")
    print(f"   {total_dl} fotos descarregadas")
    print(f"   {total_skip} já existiam (saltadas)")
    print(f"   {errors} erros")
    print(f"   Pasta: {os.path.abspath(FOTOS_DIR)}/")


if __name__ == "__main__":
    main()
