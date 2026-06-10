#!/usr/bin/env python3
"""
Crawler — douroetamega.pt
BFS completo com Playwright (browser real, executa JavaScript).
Descobre todas as páginas internas e exporta para Excel.

Instalar dependências:
    pip install playwright openpyxl beautifulsoup4
    playwright install chromium

Correr:
    python douroetamega_crawler.py

Ficheiro gerado: douroetamega_dados.xlsx  (Dados | Presença | Resumo)
"""

import asyncio
import json
import re
from collections import Counter, deque
from urllib.parse import urljoin, urlparse

from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from playwright.async_api import async_playwright, Page, BrowserContext

BASE_URL  = "https://www.douroetamega.pt"
OUTPUT    = "douroetamega_dados.xlsx"
DELAY     = 0.7          # segundos entre páginas
MAX_PAGES = 10000        # limite de segurança

# Extensões/paths a ignorar
_SKIP = [
    ".pdf", ".jpg", ".jpeg", ".png", ".gif", ".svg", ".webp",
    ".zip", ".rar", ".doc", ".docx", ".xls", ".xlsx",
    ".mp3", ".mp4", ".avi", ".mov",
    "/wp-admin/", "/wp-login", "/feed/", "/xmlrpc",
    "/cart/", "/checkout/", "/my-account/",
    "mailto:", "tel:", "javascript:", "#",
]

# Padrões que indicam uma página de conteúdo (não listagem)
_CONTENT_HINTS = re.compile(
    r"/(\d{4,}|[a-z0-9-]{10,})/?$"          # URL com ID ou slug longo no fim
    r"|/(?:poi|event|tour|lugar|local|artigo|noticia|posto|alojamento|restaurante"
    r"|gastronomia|municipio|parish|agenda|percurso|rota|trail|hiking"
    r"|accommodation|restaurant|attraction|place|point-of-interest"
    r"|ficha|detalhe|detail|item|page|post|entry)/",
    re.I
)


# ── Utilitários ───────────────────────────────────────────────────────────────

def _normalize(url: str) -> str:
    try:
        p = urlparse(url)
    except Exception:
        return ""
    netloc = (p.netloc or "").replace("www.", "")
    base_n = BASE_URL.split("//")[-1].replace("www.", "")
    if p.netloc and netloc != base_n:
        return ""
    if p.scheme and p.scheme not in ("http", "https"):
        return ""
    path = p.path or "/"
    if not path.endswith("/"):
        path += "/"
    # Preservar paginação
    q = ""
    m = re.search(r"(?:page|pag|p)=(\d+)", p.query or "")
    if m and int(m.group(1)) > 1:
        q = f"?{p.query}"
    return f"{BASE_URL}{path}{q}"


def _is_skip(url: str) -> bool:
    low = url.lower()
    return any(s in low for s in _SKIP)


def _is_content_page(url: str) -> bool:
    """Heurística: será uma página de conteúdo (item) e não uma listagem?"""
    path = urlparse(url).path
    depth = len([p for p in path.split("/") if p])
    # Páginas a profundidade >= 2 com slug ou ID tendem a ser items
    return depth >= 2 and bool(_CONTENT_HINTS.search(path))


# ── Extracção de dados ────────────────────────────────────────────────────────

def _extract_data(url: str, html: str) -> dict:
    """Extrai todos os dados disponíveis de uma página."""
    soup = BeautifulSoup(html, "html.parser")
    row: dict = {"url": url}

    # Campos da URL
    parts = url.rstrip("/").split("/")
    row["slug"]       = parts[-1] if parts[-1] else (parts[-2] if len(parts) >= 2 else "")
    row["profundidade"]= str(len([p for p in urlparse(url).path.split("/") if p]))

    # Tentar inferir categoria/tipo a partir do URL
    path_parts = [p for p in urlparse(url).path.split("/") if p]
    row["secao"]     = path_parts[0] if path_parts else ""
    row["categoria"] = path_parts[1] if len(path_parts) >= 2 else ""

    # JSON-LD (mais fiável)
    for script in soup.find_all("script", type="application/ld+json"):
        try:
            data = json.loads(script.string or "")
            if isinstance(data, list):
                data = data[0]
            if not isinstance(data, dict):
                continue
            row["nome"]      = data.get("name", "")
            row["descricao"] = data.get("description", "")
            row["tipo"]      = data.get("@type", "")
            row["website"]   = data.get("url", "")
            row["telefone"]  = data.get("telephone", "")
            row["email"]     = data.get("email", "")
            addr = data.get("address", {})
            if isinstance(addr, dict):
                row["morada"]     = addr.get("streetAddress", "")
                row["localidade"] = addr.get("addressLocality", "")
                row["regiao"]     = addr.get("addressRegion", "")
                row["pais"]       = addr.get("addressCountry", "")
            elif isinstance(addr, str):
                row["morada"] = addr
            geo = data.get("geo", {})
            if isinstance(geo, dict):
                row["latitude"]  = geo.get("latitude", "")
                row["longitude"] = geo.get("longitude", "")
            row["preco"]         = str(data.get("priceRange", "") or data.get("price", ""))
            row["horario"]       = str(data.get("openingHours", "") or "")
            row["imagem_jsonld"] = str(data.get("image", ""))
            # Datas (eventos)
            row["data_inicio"]   = str(data.get("startDate", ""))
            row["data_fim"]      = str(data.get("endDate", ""))
            row["local_evento"]  = str(data.get("location", ""))
            break
        except Exception:
            pass

    # Open Graph / meta tags
    for tag in soup.find_all("meta"):
        prop = tag.get("property") or tag.get("name") or ""
        val  = (tag.get("content") or "").strip()
        if not val:
            continue
        if prop == "og:title" and not row.get("nome"):
            row["nome"] = val
        elif prop in ("og:description", "description") and not row.get("descricao"):
            row["descricao"] = val
        elif prop == "og:image":
            row["og_image"] = val
        elif prop == "og:type":
            row["og_tipo"] = val
        elif prop == "article:published_time":
            row["data_publicacao"] = val
        elif prop == "article:modified_time":
            row["data_modificacao"] = val

    # H1
    h1 = soup.find("h1")
    if h1:
        row["h1"] = h1.get_text(" ", strip=True)
        if not row.get("nome"):
            row["nome"] = row["h1"]

    # Título da página
    title = soup.find("title")
    if title and not row.get("nome"):
        row["nome"] = title.get_text(strip=True).split("|")[0].strip()

    # Pares label → valor (dl/dt/dd, tabelas, fields)
    pares = []
    for dl in soup.find_all("dl"):
        for dt, dd in zip(dl.find_all("dt"), dl.find_all("dd")):
            t1 = dt.get_text(" ", strip=True)
            t2 = dd.get_text(" ", strip=True)
            if t1 and t2 and t1 != t2:
                pares.append((t1, t2))
    for tr in soup.find_all("tr"):
        for th, td in zip(tr.find_all("th"), tr.find_all("td")):
            t1 = th.get_text(" ", strip=True)
            t2 = td.get_text(" ", strip=True)
            if t1 and t2 and t1 != t2:
                pares.append((t1, t2))
    for el in soup.find_all(class_=re.compile(r"field|detail|info|meta|attr|prop|label", re.I)):
        lbl = el.find(["label","strong","b","dt","th","span"],
                      class_=re.compile(r"label|key|title|name", re.I))
        val = el.find(["span","p","div","dd","td"],
                      class_=re.compile(r"value|content|data|text|body", re.I))
        if lbl and val:
            t1 = lbl.get_text(" ", strip=True)
            t2 = val.get_text(" ", strip=True)
            if t1 and t2 and t1 != t2:
                pares.append((t1, t2))

    for lbl, val in pares:
        key = "campo_" + re.sub(r"[^\w]", "_", lbl.lower().strip("_"))[:40]
        if key not in row:
            row[key] = val

    # Imagens
    imgs = []
    for img in soup.find_all("img", src=True):
        src = img["src"]
        if any(x in src.lower() for x in ["/upload", "/media", "/photo", "/image",
                                            "/content", "/assets", "/files", "/img"]):
            full = urljoin(BASE_URL, src)
            if full not in imgs:
                imgs.append(full)
    row["imagens"] = " | ".join(imgs[:10])

    # Descrição longa (fallback)
    if not row.get("descricao"):
        for tag in soup.find_all(["p","div"],
                                  class_=re.compile(r"desc|content|body|text|intro|summary|about", re.I)):
            txt = tag.get_text(" ", strip=True)
            if len(txt) > 100:
                row["descricao"] = txt[:3000]
                break

    # Tags / categorias
    tags = []
    for el in soup.find_all(class_=re.compile(r"\btag\b|\bbadge\b|\bchip\b|\bcategory\b|\bcategoria\b", re.I)):
        t = el.get_text(" ", strip=True)
        if t and len(t) < 80:
            tags.append(t)
    row["tags"] = " | ".join(dict.fromkeys(tags))

    # Redes sociais
    sociais = []
    for a in soup.find_all("a", href=True):
        href = a["href"]
        if any(s in href for s in ["facebook.com", "instagram.com", "twitter.com",
                                    "youtube.com", "tiktok.com", "linkedin.com",
                                    "x.com/", "threads.net"]):
            if href not in sociais:
                sociais.append(href)
    row["redes_sociais"] = " | ".join(sociais)

    # Município (comum em sites de turismo regional)
    for pat in [r"munic[íi]pio[:\s]+([A-ZÀ-Ú][^\n<]{2,40})",
                r"concelho[:\s]+([A-ZÀ-Ú][^\n<]{2,40})"]:
        m = re.search(pat, html, re.I)
        if m and not row.get("municipio"):
            row["municipio"] = m.group(1).strip()

    return row


# ── BFS com Playwright ────────────────────────────────────────────────────────

async def crawl_all(context: BrowserContext) -> list[dict]:
    """
    BFS completo: visita todas as páginas internas.
    Para páginas que parecem conteúdo, extrai dados.
    """
    queue     = deque([BASE_URL + "/"])
    visited   = set()
    content_pages: list[dict] = []
    nav_count = 0

    page = await context.new_page()

    print(f"[BFS] A iniciar em {BASE_URL}")

    while queue and nav_count < MAX_PAGES:
        url  = queue.popleft()
        norm = _normalize(url)
        if not norm or norm in visited or _is_skip(norm):
            continue
        visited.add(norm)
        nav_count += 1

        if nav_count % 20 == 0 or nav_count <= 5:
            print(f"  [BFS {nav_count:5}] {norm}  |  items: {len(content_pages)}")

        try:
            resp = await page.goto(norm, wait_until="domcontentloaded", timeout=30000)
            if not resp or resp.status >= 400:
                continue
            await page.wait_for_timeout(800)
            html = await page.content()
        except Exception as e:
            print(f"  [ERRO] {norm}: {e}")
            await asyncio.sleep(1)
            continue

        # Extrair dados se parecer uma página de conteúdo
        if _is_content_page(norm):
            row = _extract_data(norm, html)
            if row.get("nome") or row.get("h1") or row.get("descricao"):
                content_pages.append(row)
                if len(content_pages) % 100 == 0:
                    print(f"  [Items] {len(content_pages)} páginas de conteúdo extraídas")

        # Seguir todos os links internos
        soup = BeautifulSoup(html, "html.parser")
        for a in soup.find_all("a", href=True):
            href = a["href"].strip()
            if not href or _is_skip(href):
                continue
            full  = urljoin(norm, href).split("#")[0]
            norm2 = _normalize(full)
            if norm2 and norm2 not in visited and not _is_skip(norm2):
                queue.append(norm2)

        await asyncio.sleep(DELAY)

    await page.close()
    print(f"\n[BFS] {nav_count} páginas visitadas | {len(content_pages)} items extraídos")
    return content_pages


# ── Excel ─────────────────────────────────────────────────────────────────────

_COLS_PRIORITY = [
    "nome", "h1", "secao", "categoria", "slug", "tipo",
    "descricao", "morada", "localidade", "municipio", "regiao", "pais",
    "latitude", "longitude", "telefone", "email", "website",
    "preco", "horario", "data_inicio", "data_fim", "local_evento",
    "data_publicacao", "imagem_jsonld", "og_image", "og_tipo",
    "imagens", "tags", "redes_sociais", "profundidade", "url",
]

_HDR_FILL = PatternFill("solid", fgColor="17375E")
_HDR_FONT = Font(color="FFFFFF", bold=True)
_X_FILL   = PatternFill("solid", fgColor="E2EFDA")


def _write_sheet(ws, headers, rows, presence=False):
    ws.append(headers)
    for i in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=i)
        c.fill = _HDR_FILL
        c.font = _HDR_FONT
        c.alignment = Alignment(horizontal="center")
    for row in rows:
        vals = []
        for col in headers:
            v = str(row.get(col, "") or "").strip()
            vals.append("X" if (presence and v) else ("" if presence else v))
        ws.append(vals)
    if presence:
        for r in ws.iter_rows(min_row=2):
            for cell in r:
                if cell.value == "X":
                    cell.fill = _X_FILL
                    cell.alignment = Alignment(horizontal="center")
    for i, name in enumerate(headers, 1):
        mx = len(name)
        for r in ws.iter_rows(min_row=2, min_col=i, max_col=i):
            for cell in r:
                if cell.value:
                    mx = max(mx, min(len(str(cell.value)), 60))
        ws.column_dimensions[get_column_letter(i)].width = mx + 2


def save_excel(dados, output=OUTPUT):
    if not dados:
        print("[Erro] Sem dados para guardar.")
        return

    all_keys, seen = [], set()
    for row in dados:
        for k in row:
            if k not in seen:
                seen.add(k)
                all_keys.append(k)

    priority = [c for c in _COLS_PRIORITY if c in seen]
    headers  = priority + [c for c in all_keys if c not in priority]

    wb = Workbook()

    ws = wb.active
    ws.title = "Dados"
    _write_sheet(ws, headers, dados)

    ws2 = wb.create_sheet("Presença")
    _write_sheet(ws2, headers, dados, presence=True)

    ws3 = wb.create_sheet("Resumo")
    ws3.append(["secao", "categoria", "total"])
    for i in range(1, 4):
        c = ws3.cell(row=1, column=i)
        c.fill = _HDR_FILL; c.font = _HDR_FONT
        c.alignment = Alignment(horizontal="center")
    counter = Counter(
        (str(r.get("secao", "")), str(r.get("categoria", ""))) for r in dados
    )
    for (s, c), n in sorted(counter.items()):
        ws3.append([s, c, n])
    for i in range(1, 4):
        ws3.column_dimensions[get_column_letter(i)].width = 28

    wb.save(output)
    print(f"\n✅  {output}")
    print(f"    {len(dados)} linhas × {len(headers)} colunas")
    print(f"    Sheets: Dados | Presença (X/vazio) | Resumo por secção")


# ── Main ──────────────────────────────────────────────────────────────────────

async def main():
    print("=" * 60)
    print("  douroetamega.pt — Crawler (Playwright BFS)")
    print("=" * 60)

    async with async_playwright() as p:
        try:
            browser = await p.chromium.launch(headless=True)
        except Exception:
            try:
                browser = await p.chromium.launch(
                    headless=True, executable_path="/usr/bin/chromium-browser"
                )
            except Exception:
                browser = await p.chromium.launch(
                    headless=True, executable_path="/usr/bin/chromium"
                )

        context = await browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                       "AppleWebKit/537.36 (KHTML, like Gecko) "
                       "Chrome/136.0.0.0 Safari/537.36",
            locale="pt-PT",
        )

        # Warm-up
        pg = await context.new_page()
        try:
            await pg.goto(BASE_URL + "/", wait_until="domcontentloaded", timeout=30000)
            await pg.wait_for_timeout(1500)
            print(f"[Session] {BASE_URL} acessível\n")
        except Exception as e:
            print(f"[Session] {e}\n")
        await pg.close()

        # Crawl BFS completo
        dados = await crawl_all(context)

        # Guardas parciais a cada 500
        if dados:
            for i in range(500, len(dados), 500):
                save_excel(dados[:i], OUTPUT.replace(".xlsx", f"_parcial_{i}.xlsx"))

        await browser.close()

    save_excel(dados)


if __name__ == "__main__":
    asyncio.run(main())
