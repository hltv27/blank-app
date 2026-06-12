#!/usr/bin/env python3
"""
Crawler — turismo.douroetamega.pt + aboboreira.douroetamega.pt  (Selenium)
Abordagem igual ao guidedbynature: percorre cada categoria página a página.

Instalar:
    pkg install chromium
    pip install selenium openpyxl beautifulsoup4

Correr:
    python douroetamega_crawler_selenium.py
"""

import json
import os
import re
import shutil
import time
from collections import Counter
from urllib.parse import urljoin, urlparse

from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.common.exceptions import TimeoutException, WebDriverException

# ── Configuração ──────────────────────────────────────────────────────────────

TURISMO_URL    = "https://turismo.douroetamega.pt"
ABOBOREIRA_URL = "https://aboboreira.douroetamega.pt"
OUTPUT         = "douroetamega_dados.xlsx"
DELAY          = 0.8   # segundos entre páginas de item

# Categorias do turismo.douroetamega.pt
# Cada entrada é o path da página de listagem (sem domínio)
CATEGORIAS_TURISMO = [
    "o-que-ver/patrimonio",
    "o-que-ver/postos-de-turismo",
    "o-que-ver/miradouros-e-vistas",
    "o-que-ver/espacos-verdes",
    "o-que-fazer/cultura-e-arte",
    "o-que-fazer/museus",
    "o-que-fazer/artesanato",
    "o-que-fazer/comercializacao",
    "o-que-fazer/animacao-cultural-recreativa-e-de-lazer",
    "o-que-fazer/agentes-culturais",
    "o-que-fazer/congressos-e-exposicoes",
    "o-que-fazer/desporto-e-lazer",
    "o-que-fazer/empresas-de-animacao-turistica",
    "o-que-fazer/aldeias-de-portugal",
    "o-que-fazer/rota-do-romanico",
    "o-que-fazer/rotas-e-percursos/percursos-pedestres",
    "o-que-fazer/rotas-e-percursos/btt",
    "o-que-fazer/rotas-e-percursos/roteiros-baixo-tamega",
    "o-que-fazer/rotas-e-percursos/outros-roteiros",
    "o-que-fazer/rotas-e-percursos/serra-da-aboboreira",
    "o-que-fazer/escapadinhas",
    "o-que-fazer/verde-sentido",
    "o-que-fazer/caves",
    "onde-dormir/turismo-rural",
    "onde-dormir/turismo-de-habitacao",
    "onde-dormir/alojamento-local",
    "onde-dormir/albergues-abrigos-e-pousadas",
    "onde-dormir/parques-de-campismo",
    "onde-dormir/hoteis",
    "onde-comer",
    "agenda/eventos",
    "pages/856",   # Aldeias de Portugal
]

# Categorias do aboboreira.douroetamega.pt
CATEGORIAS_ABOBOREIRA = [
    "pages/1008",            # percursos pedestres
    "rotas-e-percursos",
    "paisagem-protegida-regional",
]

# Regex para encontrar URLs de items nas páginas de listagem
# Captura: /geo_artigo[-N]/slug  ou  /percurso/slug  ou  ?geo_article_id=N
_ITEM_RE = re.compile(
    r'href=["\']('
    r'[^"\']*geo_artigo(?:-\d+)?/[^"\'/?][^"\'/?]*'  # geo_artigo/slug
    r'|[^"\']*?/percurso/[^"\'/?][^"\'/?]*'           # /percurso/slug
    r'|[^"\']*?/evento/[^"\'/?][^"\'/?]*'             # /evento/slug
    r'|[^"\']*?[?&]geo_article_id=\d+'                # ?geo_article_id=N
    r')["\']',
    re.I
)


# ── Selenium ──────────────────────────────────────────────────────────────────

def _make_driver() -> webdriver.Chrome:
    opts = Options()
    opts.add_argument("--headless")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--disable-gpu")
    opts.add_argument("--window-size=1280,800")
    opts.add_argument("--lang=pt-PT")
    opts.add_argument(
        "--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/136.0.0.0 Safari/537.36"
    )
    chromium_candidates = [
        "/data/data/com.termux/files/usr/bin/chromium-browser",
        "/data/data/com.termux/files/usr/bin/chromium",
        "/usr/bin/chromium-browser",
        "/usr/bin/chromium",
        shutil.which("chromium-browser") or "",
        shutil.which("chromium") or "",
    ]
    chromium_bin = next((p for p in chromium_candidates if p and os.path.exists(p)), None)
    if chromium_bin:
        opts.binary_location = chromium_bin

    driver_candidates = [
        "/data/data/com.termux/files/usr/bin/chromedriver",
        "/usr/bin/chromedriver",
        "/usr/local/bin/chromedriver",
        shutil.which("chromedriver") or "",
    ]
    driver_bin = next((p for p in driver_candidates if p and os.path.exists(p)), None)
    if not driver_bin:
        raise RuntimeError("chromedriver não encontrado. Instala: pkg install chromium")

    svc = Service(executable_path=driver_bin)
    return webdriver.Chrome(service=svc, options=opts)


def _get_listing(driver: webdriver.Chrome, url: str) -> str | None:
    """Carrega uma página de listagem: espera JS + scroll para lazy loading."""
    try:
        driver.get(url)
        time.sleep(2.5)
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
        time.sleep(1.5)
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
        time.sleep(1.0)
        return driver.page_source
    except (TimeoutException, WebDriverException) as e:
        print(f"  [ERRO listing] {url}: {e}")
        return None


def _get_item(driver: webdriver.Chrome, url: str) -> str | None:
    """Carrega uma página de item: espera standard."""
    try:
        driver.get(url)
        time.sleep(1.5)
        return driver.page_source
    except (TimeoutException, WebDriverException) as e:
        print(f"  [ERRO item] {url}: {e}")
        return None


# ── FASE 1: Descoberta por categoria ─────────────────────────────────────────

def _extract_item_urls(html: str, base: str) -> set[str]:
    """Extrai URLs de items de uma página de listagem."""
    urls = set()
    for m in _ITEM_RE.finditer(html):
        href = m.group(1).strip()
        if not href:
            continue
        full = urljoin(base, href).split("#")[0].rstrip("&")
        # Normaliza trailing slash
        p = urlparse(full)
        if p.query:
            full = f"{p.scheme}://{p.netloc}{p.path.rstrip('/')}?{p.query}"
        else:
            if not p.path.endswith("/"):
                full += "/"
        urls.add(full)
    return urls


def discover_urls(driver: webdriver.Chrome) -> list[str]:
    """
    Visita cada categoria página a página (igual ao guidedbynature).
    Usa scroll para activar lazy loading dos items.
    """
    all_urls: set[str] = set()

    entries = (
        [(TURISMO_URL, cat) for cat in CATEGORIAS_TURISMO] +
        [(ABOBOREIRA_URL, cat) for cat in CATEGORIAS_ABOBOREIRA]
    )

    for base_url, cat in entries:
        cat_base = f"{base_url}/{cat}/"
        cat_found = 0

        for pg in range(1, 300):
            url = cat_base if pg == 1 else f"{cat_base}?page={pg}"
            html = _get_listing(driver, url)
            if not html:
                break

            # Verifica se a página existe (404 / redirecciona para homepage)
            if pg > 1:
                soup_check = BeautifulSoup(html, "html.parser")
                title = (soup_check.find("title") or soup_check.find("h1") or "")
                title_txt = title.get_text(strip=True) if title else ""
                # Se o título for igual à homepage ou a página for vazia, parar
                if "AMBT Turismo" == title_txt or not title_txt:
                    break

            new = _extract_item_urls(html, base_url)
            before = len(all_urls)
            all_urls |= new
            added = len(all_urls) - before
            cat_found += added

            if added == 0:
                break   # sem novos items → última página desta categoria

        if cat_found:
            print(f"  [{cat}] {cat_found} items")

    print(f"\n  Total: {len(all_urls)} URLs únicos\n")
    return sorted(all_urls)


# ── FASE 2: Extracção ─────────────────────────────────────────────────────────

def extract_page(url: str, html: str) -> dict:
    soup = BeautifulSoup(html, "html.parser")
    parsed = urlparse(url)
    row: dict = {
        "url":     url,
        "dominio": parsed.netloc,
    }

    path_parts = [p for p in parsed.path.split("/") if p]
    _markers = {"geo_artigo", "percurso", "evento"}
    marker_idx = next(
        (i for i, p in enumerate(path_parts)
         if p.startswith("geo_artigo") or p in _markers),
        len(path_parts)
    )
    row["secao"]        = path_parts[0] if path_parts else ""
    row["categoria"]    = path_parts[1] if len(path_parts) >= 2 else ""
    row["subcategoria"] = path_parts[2] if len(path_parts) >= 3 and marker_idx > 2 else ""
    last = path_parts[-1] if path_parts else ""
    row["slug"] = last if not last.isdigit() else (path_parts[-2] if len(path_parts) >= 2 else "")
    row["id"]   = ""
    if parsed.query:
        m = re.search(r"geo_article_id=(\d+)", parsed.query)
        if m:
            row["id"] = m.group(1)
    row["profundidade"] = str(len(path_parts))

    # JSON-LD
    for script in soup.find_all("script", type="application/ld+json"):
        try:
            data = json.loads(script.string or "")
            if isinstance(data, list):
                data = data[0]
            if not isinstance(data, dict):
                continue
            row["nome"]       = data.get("name", "")
            row["descricao"]  = data.get("description", "")
            row["tipo"]       = data.get("@type", "")
            row["website"]    = data.get("url", "")
            row["telefone"]   = data.get("telephone", "")
            row["email"]      = data.get("email", "")
            addr = data.get("address", {})
            if isinstance(addr, dict):
                row["morada"]     = addr.get("streetAddress", "")
                row["localidade"] = addr.get("addressLocality", "")
                row["regiao"]     = addr.get("addressRegion", "")
                row["pais"]       = addr.get("addressCountry", "")
            elif isinstance(addr, str):
                row["morada"] = addr
            geo = data.get("geo", {})
            if isinstance(geo, dict):
                row["latitude"]  = geo.get("latitude", "")
                row["longitude"] = geo.get("longitude", "")
            row["preco"]         = str(data.get("priceRange", "") or data.get("price", ""))
            row["horario"]       = str(data.get("openingHours", "") or "")
            row["imagem_jsonld"] = str(data.get("image", ""))
            row["data_inicio"]   = str(data.get("startDate", ""))
            row["data_fim"]      = str(data.get("endDate", ""))
            row["local_evento"]  = str(data.get("location", ""))
            for src_k, dst_k in [("distance","distancia"),("length","distancia"),
                                   ("elevation","elevacao"),("ascent","elevacao"),
                                   ("difficulty","dificuldade"),("duration","duracao")]:
                if data.get(src_k) and not row.get(dst_k):
                    row[dst_k] = str(data[src_k])
            break
        except Exception:
            pass

    # Open Graph / meta
    for tag in soup.find_all("meta"):
        prop = tag.get("property") or tag.get("name") or ""
        val  = (tag.get("content") or "").strip()
        if not val:
            continue
        if prop == "og:title" and not row.get("nome"):
            row["nome"] = val
        elif prop in ("og:description","description") and not row.get("descricao"):
            row["descricao"] = val
        elif prop == "og:image":
            row["og_image"] = val
        elif prop == "og:type":
            row["og_tipo"] = val
        elif prop == "article:published_time":
            row["data_publicacao"] = val
        elif prop == "article:modified_time":
            row["data_modificacao"] = val
        elif prop == "keywords":
            row["keywords"] = val

    h1 = soup.find("h1")
    if h1:
        row["h1"] = h1.get_text(" ", strip=True)
        if not row.get("nome"):
            row["nome"] = row["h1"]
    if not row.get("nome"):
        h2 = soup.find("h2")
        if h2:
            row["nome"] = h2.get_text(" ", strip=True)
    title = soup.find("title")
    if title and not row.get("nome"):
        row["nome"] = title.get_text(strip=True).split("|")[0].split("–")[0].strip()

    # dl/dt/dd + tabelas
    pares: list[tuple] = []
    for dl in soup.find_all("dl"):
        for dt, dd in zip(dl.find_all("dt"), dl.find_all("dd")):
            t1 = dt.get_text(" ", strip=True)
            t2 = dd.get_text(" ", strip=True)
            if t1 and t2 and t1 != t2:
                pares.append((t1, t2))
    for tr in soup.find_all("tr"):
        for th, td in zip(tr.find_all("th"), tr.find_all("td")):
            t1 = th.get_text(" ", strip=True)
            t2 = td.get_text(" ", strip=True)
            if t1 and t2 and t1 != t2:
                pares.append((t1, t2))

    _label_map = {
        "municipio":     ["município","municipio","concelho"],
        "morada":        ["morada","endereço","address","localização"],
        "telefone":      ["telefone","telef","tel.","contacto"],
        "email":         ["e-mail","email","correio"],
        "horario":       ["horário","horario","horas","schedule","funcionamento"],
        "website":       ["website","site","página web"],
        "preco":         ["preço","entrada","admissão","bilhete"],
        "latitude":      ["latitude","lat"],
        "longitude":     ["longitude","lon","lng"],
        "distancia":     ["distância","distancia","comprimento","extensão","length"],
        "elevacao":      ["elevação","elevacao","desnível","desnivel","altitude"],
        "dificuldade":   ["dificuldade","difficulty","nível","grau"],
        "duracao":       ["duração","duracao","duration","tempo estimado"],
        "acessos":       ["acesso","acessos","como chegar"],
        "classificacao": ["classificação","tipo de percurso","tipologia"],
        "capacidade":    ["capacidade","lugares","camas","quartos"],
    }
    for lbl_raw, val_raw in pares:
        lbl_n = lbl_raw.lower().strip()
        matched = False
        for field, syns in _label_map.items():
            if any(s in lbl_n for s in syns):
                if not row.get(field):
                    row[field] = val_raw
                matched = True
                break
        if not matched:
            key = "campo_" + re.sub(r"[^\w]", "_", lbl_n.strip("_"))[:40]
            if key not in row:
                row[key] = val_raw

    # Imagens — apenas do domínio douroetamega.pt
    _JUNK = ("googlelogo","sunny.png","weather","spinner","loading",
             "placeholder","favicon","maps.gstatic","gstatic.com",
             "googleapis.com","icon-","/icons/","/logo")
    imgs = []
    for img in soup.find_all("img", src=True):
        src = img["src"].strip()
        if not src:
            continue
        full = urljoin(url, src)
        if "douroetamega.pt" not in full:
            continue
        if any(j in full.lower() for j in _JUNK):
            continue
        if full not in imgs:
            imgs.append(full)
    row["imagens"] = " | ".join(imgs[:20])

    # Descrição
    if not row.get("descricao"):
        for tag in soup.find_all(["div","article","section"],
                                  class_=re.compile(
                                      r"desc|content|body|text|intro|summary"
                                      r"|about|corpo|conteudo|article|detail"
                                      r"|ficha|info|main|detalhe", re.I)):
            cls = " ".join(tag.get("class", []))
            if re.search(r"nav|menu|sidebar|header|footer|bread", cls, re.I):
                continue
            txt = tag.get_text(" ", strip=True)
            if len(txt) > 120:
                row["descricao"] = txt[:3000]
                break
    if not row.get("descricao"):
        for tag in soup.find_all("p"):
            if tag.parent and tag.parent.name in ("nav","header","footer"):
                continue
            txt = tag.get_text(" ", strip=True)
            if len(txt) > 80:
                row["descricao"] = txt[:3000]
                break

    # Tags
    tags = []
    for el in soup.find_all(class_=re.compile(r"\btag\b|\bbadge\b|\bchip\b|\bcategoria\b", re.I)):
        t = el.get_text(" ", strip=True)
        if t and len(t) < 80:
            tags.append(t)
    row["tags"] = " | ".join(dict.fromkeys(tags))

    # Redes sociais
    sociais = []
    for a in soup.find_all("a", href=True):
        href = a["href"]
        if any(s in href for s in ["facebook.com","instagram.com","twitter.com",
                                    "youtube.com","tiktok.com","linkedin.com","x.com/"]):
            if href not in sociais:
                sociais.append(href)
    row["redes_sociais"] = " | ".join(sociais)

    # Regex no texto plano
    raw = soup.get_text(" ")
    for pat in [r"munic[íi]pio[:\s]+([A-ZÀ-Úa-zà-ú][^\n.,;]{2,40})",
                r"concelho[:\s]+([A-ZÀ-Úa-zà-ú][^\n.,;]{2,40})"]:
        m = re.search(pat, raw, re.I)
        if m and not row.get("municipio"):
            row["municipio"] = m.group(1).strip()
    if not row.get("distancia"):
        m = re.search(r"(\d+(?:[.,]\d+)?\s*km)", raw, re.I)
        if m:
            row["distancia"] = m.group(1).strip()
    if not row.get("elevacao"):
        m = re.search(r"desnível\s*[:\s]+(\d+\s*m)|(\d+)\s*m\s*(?:de\s+)?desnível", raw, re.I)
        if m:
            row["elevacao"] = (m.group(1) or m.group(2) or "").strip() + " m"
    if not row.get("dificuldade"):
        m = re.search(r"dificuldade[:\s]+([^\n.,;]{3,30})", raw, re.I)
        if m:
            row["dificuldade"] = m.group(1).strip()
    if not row.get("duracao"):
        m = re.search(r"dura[çc][aã]o[:\s]+([\dhHmM: ]+)|(\d+h\d*(?:min)?|\d+\s*hora[s]?)", raw, re.I)
        if m:
            row["duracao"] = (m.group(1) or m.group(2) or "").strip()

    return row


# ── Excel ─────────────────────────────────────────────────────────────────────

_COLS_PRIORITY = [
    "id","nome","h1","dominio","secao","categoria","subcategoria","slug","tipo",
    "descricao","municipio","morada","localidade","regiao","pais",
    "latitude","longitude","telefone","email","website",
    "preco","horario","capacidade",
    "distancia","elevacao","dificuldade","duracao","classificacao","acessos",
    "data_inicio","data_fim","local_evento",
    "data_publicacao","data_modificacao","keywords",
    "imagem_jsonld","og_image","og_tipo",
    "imagens","tags","redes_sociais",
    "profundidade","url",
]

_HDR_FILL = PatternFill("solid", fgColor="17375E")
_HDR_FONT = Font(color="FFFFFF", bold=True)
_X_FILL   = PatternFill("solid", fgColor="E2EFDA")

# Caracteres ilegais para células Excel (controlo ASCII exceto tab/LF/CR)
_ILLEGAL_RE = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]')

def _clean(v: str) -> str:
    return _ILLEGAL_RE.sub('', str(v or "")).strip()


# Colunas que na sheet Presença mostram o valor real (não X/branco)
_ID_COLS = {"id","nome","dominio","secao","categoria","subcategoria","slug","url"}

def _write_sheet(ws, headers, rows, presence=False):
    ws.append(headers)
    for i in range(1, len(headers)+1):
        c = ws.cell(row=1, column=i)
        c.fill = _HDR_FILL; c.font = _HDR_FONT
        c.alignment = Alignment(horizontal="center")
    for row in rows:
        vals = []
        for col in headers:
            v = _clean(row.get(col, ""))
            if presence:
                if col in _ID_COLS:
                    vals.append(v)          # mostra valor real
                else:
                    vals.append("X" if v else "")
            else:
                vals.append(v)
        ws.append(vals)
    if presence:
        for r in ws.iter_rows(min_row=2):
            for cell in r:
                if cell.value == "X":
                    cell.fill = _X_FILL
                    cell.alignment = Alignment(horizontal="center")
    for i, name in enumerate(headers, 1):
        mx = len(name)
        for r in ws.iter_rows(min_row=2, min_col=i, max_col=i):
            for cell in r:
                if cell.value:
                    mx = max(mx, min(len(str(cell.value)), 60))
        ws.column_dimensions[get_column_letter(i)].width = mx + 2


def save_excel(dados, output=OUTPUT):
    if not dados:
        print("[Erro] Sem dados.")
        return
    all_keys, seen = [], set()
    for row in dados:
        for k in row:
            if k not in seen:
                seen.add(k); all_keys.append(k)
    priority = [c for c in _COLS_PRIORITY if c in seen]
    headers  = priority + [c for c in all_keys if c not in priority]

    wb = Workbook()
    ws = wb.active; ws.title = "Dados"
    _write_sheet(ws, headers, dados)
    ws2 = wb.create_sheet("Presença")
    _write_sheet(ws2, headers, dados, presence=True)
    ws3 = wb.create_sheet("Resumo")
    ws3.append(["dominio","secao","categoria","total"])
    for i in range(1, 5):
        c = ws3.cell(row=1, column=i)
        c.fill = _HDR_FILL; c.font = _HDR_FONT
        c.alignment = Alignment(horizontal="center")
    counter = Counter(
        (str(r.get("dominio","")), str(r.get("secao","")), str(r.get("categoria","")))
        for r in dados
    )
    for (d, s, c), n in sorted(counter.items()):
        ws3.append([d, s, c, n])
    for i in range(1, 5):
        ws3.column_dimensions[get_column_letter(i)].width = 30
    wb.save(output)
    print(f"\n✅  {output}  —  {len(dados)} linhas × {len(headers)} colunas")
    print(f"    Sheets: Dados | Presença | Resumo")


# ── Main ──────────────────────────────────────────────────────────────────────

URLS_FILE = "douroetamega_urls.txt"


def _phase2(urls: list[str], batch_size: int = 50) -> list[dict]:
    """Extracção Selenium em batches — reinicia browser a cada batch para evitar OOM."""
    total = len(urls)
    dados = []
    for batch_start in range(0, total, batch_size):
        batch = urls[batch_start:batch_start+batch_size]
        print(f"\n  [Batch] {batch_start+1}–{batch_start+len(batch)} / {total}")
        driver = _make_driver()
        try:
            for j, url in enumerate(batch, 1):
                i = batch_start + j
                if i <= 10 or i % 50 == 0:
                    slug = url.rstrip("/").split("/")[-1].split("?")[0]
                    print(f"  [{i:5}/{total}] {slug[:60]}")
                html = _get_item(driver, url)
                if html:
                    dados.append(extract_page(url, html))
                if i % 200 == 0 and dados:
                    save_excel(dados, OUTPUT.replace(".xlsx", f"_parcial_{i}.xlsx"))
                time.sleep(DELAY)
        finally:
            try: driver.quit()
            except Exception: pass
    return dados


def main():
    print("=" * 60)
    print("  douroetamega.pt — Crawler (turismo + aboboreira)")
    print("=" * 60)

    # ── Fase 1: Descoberta por categoria ─────────────────────────────────────
    if os.path.exists(URLS_FILE):
        with open(URLS_FILE) as f:
            urls = [l.strip() for l in f if l.strip()]
        print(f"\n[Fase 1] {len(urls)} URLs de {URLS_FILE}  (apaga para re-descobrir)\n")
    else:
        print(f"\n[Fase 1] A descobrir items por categoria…\n")
        driver = _make_driver()
        try:
            driver.get(TURISMO_URL + "/")
            time.sleep(2)
            print(f"  Session OK\n")
        except Exception as e:
            print(f"  Aviso: {e}\n")
        try:
            urls = discover_urls(driver)
        finally:
            driver.quit()

        if not urls:
            print("[Aviso] Nenhum URL encontrado.")
            return

        with open(URLS_FILE, "w") as f:
            f.write("\n".join(urls))
        print(f"  URLs guardados em {URLS_FILE}")

    # ── Fase 2: Extracção item a item ─────────────────────────────────────────
    print(f"[Fase 2] A extrair {len(urls)} items (batches de 50)…\n")
    dados = _phase2(urls, batch_size=50)
    save_excel(dados)

    # ── Fase 3: Fotos ─────────────────────────────────────────────────────────
    try:
        from download_fotos import main as dl_fotos
        print("\n[Fase 3] A descarregar fotos…")
        dl_fotos()
    except Exception as e:
        print(f"\n[Fase 3] {e}")
        print("  Corre manualmente: python download_fotos.py")


if __name__ == "__main__":
    main()
