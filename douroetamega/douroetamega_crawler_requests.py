#!/usr/bin/env python3
"""
Crawler — douroetamega.pt  (requests, duas fases)
Fase 1: BFS — segue todos os links internos e recolhe URLs de items
Fase 2: Extracção — visita cada URL e extrai dados estruturados

(As páginas de listagem são renderizadas por JS como no guidedbynature,
por isso usamos BFS em vez de paginar categorias directamente.)

Instalar:
    pip install cloudscraper openpyxl beautifulsoup4

Correr:
    python douroetamega_crawler_requests.py
"""

import json
import os
import re
import time
from collections import Counter, deque
from urllib.parse import urljoin, urlparse

try:
    import cloudscraper
    _session = cloudscraper.create_scraper()
    print("[HTTP] cloudscraper OK")
except ImportError:
    import requests
    _session = requests.Session()
    print("[HTTP] requests")

_session.headers.update({
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                  "AppleWebKit/537.36 (KHTML, like Gecko) "
                  "Chrome/136.0.0.0 Safari/537.36",
    "Accept-Language": "pt-PT,pt;q=0.9,en;q=0.8",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Referer": "https://www.douroetamega.pt/",
})

from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

BASE_URL      = "https://www.douroetamega.pt"
OUTPUT        = "douroetamega_dados.xlsx"
FOTOS_DIR     = "douroetamega_fotos"
DELAY_BFS     = 0.5
DELAY_EXT     = 0.8
DELAY_FOTO    = 0.3
MAX_BFS       = 15000

# Extensões/paths a ignorar
_SKIP_EXT  = (".pdf",".jpg",".jpeg",".png",".gif",".svg",".webp",
              ".zip",".rar",".doc",".docx",".xls",".xlsx",
              ".mp3",".mp4",".avi",".mov",".woff",".woff2",".css",".js",".ico")
_SKIP_PATH = ("/wp-admin/","/wp-login","/feed/","/xmlrpc",
              "/cart/","/checkout/","/my-account/","/search/","/tag/","/author/",
              "/admin/","/manager/","/cms/","/backend/",
              "/login","/logout","/register","/api/")


def _is_skip(url: str) -> bool:
    low = url.lower()
    return (any(low.endswith(e) for e in _SKIP_EXT) or
            any(p in low for p in _SKIP_PATH) or
            low.startswith(("mailto:","tel:","javascript:")) or
            "#" in low)


def _norm(url: str) -> str:
    """Normaliza URL: strip fragment, trailing slash, verifica domínio."""
    try:
        p = urlparse(url.split("#")[0])
    except Exception:
        return ""
    if p.scheme not in ("", "http", "https"):
        return ""
    netloc = (p.netloc or "").replace("www.", "")
    if netloc and netloc != "douroetamega.pt":
        return ""
    path = p.path or "/"
    # Ignora caminhos com extensão de ficheiro
    last = path.split("/")[-1]
    if "." in last and not last.startswith("."):
        return ""
    if not path.endswith("/"):
        path += "/"
    # Preserva paginação (?page=N, ?p=N)
    qs = ""
    if p.query:
        m = re.search(r"(?:page|p)=(\d+)", p.query)
        if m and int(m.group(1)) > 1:
            qs = f"?{p.query}"
    return f"{BASE_URL}{path}{qs}"


def _is_content_page(url: str) -> bool:
    """
    Identifica páginas de item (não homepage nem secções de topo).
    Profundidade >= 2 com pelo menos um segmento longo ou numérico.
    Exemplos que devem passar:
      /atualidade/agenda/evento/workshop-azulejo/   (depth 4)
      /pages/730/                                   (depth 2, numérico)
      /associacao/mensagem-do-presidente/           (depth 2, slug longo)
    """
    path = urlparse(url).path
    parts = [p for p in path.split("/") if p]
    if len(parts) < 2:
        return False
    # Pelo menos um segmento numérico ou slug com 5+ chars
    return any(
        p.isdigit() or len(p) >= 5
        for p in parts
    )


def _get(url: str, retries: int = 3) -> str | None:
    for attempt in range(retries):
        try:
            r = _session.get(url, timeout=25)
            if r.ok:
                return r.text
            if r.status_code in (403, 404, 410):
                return None
        except Exception as e:
            if attempt == retries - 1:
                print(f"  [ERRO] {url}: {e}")
        time.sleep(1.5 ** (attempt + 1))
    return None


# ── FASE 1: BFS — Descoberta de URLs ─────────────────────────────────────────

def discover_urls() -> list[str]:
    """
    BFS completo: segue todos os links internos.
    Recolhe apenas URLs que passam em _is_content_page().
    As páginas de listagem são JS-rendered (sem links no HTML estático),
    por isso o BFS parte da navegação que liga directamente aos items.
    """
    queue   = deque([BASE_URL + "/"])
    visited: set[str] = set()
    content: set[str] = set()
    nav_count = 0

    print(f"  A iniciar BFS em {BASE_URL}")

    while queue and nav_count < MAX_BFS:
        url  = queue.popleft()
        norm = _norm(url)
        if not norm or norm in visited or _is_skip(norm):
            continue
        visited.add(norm)
        nav_count += 1

        if nav_count % 100 == 0 or nav_count <= 5:
            print(f"  [BFS {nav_count:5}] {norm.replace(BASE_URL,'')}  |  items: {len(content)}")

        html = _get(norm)
        if not html:
            continue

        if _is_content_page(norm):
            content.add(norm)

        # Seguir todos os links internos
        soup = BeautifulSoup(html, "html.parser")
        for a in soup.find_all("a", href=True):
            href = a["href"].strip()
            if not href or _is_skip(href):
                continue
            full  = urljoin(norm, href)
            norm2 = _norm(full)
            if norm2 and norm2 not in visited:
                queue.append(norm2)

        time.sleep(DELAY_BFS)

    print(f"\n  BFS: {nav_count} páginas visitadas | {len(content)} URLs de items")
    return sorted(content)


# ── FASE 2: Extracção ─────────────────────────────────────────────────────────

def extract_page(url: str, html: str) -> dict:
    """Extrai todos os dados disponíveis de uma página de item."""
    soup = BeautifulSoup(html, "html.parser")
    row: dict = {"url": url}

    # Info do URL
    path_parts = [p for p in urlparse(url).path.split("/") if p]
    last       = path_parts[-1] if path_parts else ""
    row["id"]          = last if last.isdigit() else ""
    row["slug"]        = last if not last.isdigit() else (path_parts[-2] if len(path_parts) >= 2 else "")
    row["secao"]       = path_parts[0] if path_parts else ""
    row["categoria"]   = path_parts[1] if len(path_parts) >= 2 else ""
    row["subcategoria"]= path_parts[2] if len(path_parts) >= 3 else ""
    row["profundidade"]= str(len(path_parts))

    # JSON-LD (dados estruturados — mais fiáveis)
    for script in soup.find_all("script", type="application/ld+json"):
        try:
            data = json.loads(script.string or "")
            if isinstance(data, list):
                data = data[0]
            if not isinstance(data, dict):
                continue
            row["nome"]       = data.get("name", "")
            row["descricao"]  = data.get("description", "")
            row["tipo"]       = data.get("@type", "")
            row["website"]    = data.get("url", "")
            row["telefone"]   = data.get("telephone", "")
            row["email"]      = data.get("email", "")
            addr = data.get("address", {})
            if isinstance(addr, dict):
                row["morada"]     = addr.get("streetAddress", "")
                row["localidade"] = addr.get("addressLocality", "")
                row["regiao"]     = addr.get("addressRegion", "")
                row["pais"]       = addr.get("addressCountry", "")
            elif isinstance(addr, str):
                row["morada"] = addr
            geo = data.get("geo", {})
            if isinstance(geo, dict):
                row["latitude"]  = geo.get("latitude", "")
                row["longitude"] = geo.get("longitude", "")
            row["preco"]         = str(data.get("priceRange", "") or data.get("price", ""))
            row["horario"]       = str(data.get("openingHours", "") or "")
            row["imagem_jsonld"] = str(data.get("image", ""))
            row["data_inicio"]   = str(data.get("startDate", ""))
            row["data_fim"]      = str(data.get("endDate", ""))
            row["local_evento"]  = str(data.get("location", ""))
            break
        except Exception:
            pass

    # Open Graph / meta tags
    for tag in soup.find_all("meta"):
        prop = tag.get("property") or tag.get("name") or ""
        val  = (tag.get("content") or "").strip()
        if not val:
            continue
        if prop == "og:title" and not row.get("nome"):
            row["nome"] = val
        elif prop in ("og:description","description") and not row.get("descricao"):
            row["descricao"] = val
        elif prop == "og:image":
            row["og_image"] = val
        elif prop == "og:type":
            row["og_tipo"] = val
        elif prop == "article:published_time":
            row["data_publicacao"] = val
        elif prop == "article:modified_time":
            row["data_modificacao"] = val
        elif prop == "keywords":
            row["keywords"] = val

    # H1 → nome
    h1 = soup.find("h1")
    if h1:
        row["h1"] = h1.get_text(" ", strip=True)
        if not row.get("nome"):
            row["nome"] = row["h1"]

    # H2 fallback
    if not row.get("nome"):
        h2 = soup.find("h2")
        if h2:
            row["nome"] = h2.get_text(" ", strip=True)

    # <title> último fallback
    title = soup.find("title")
    if title and not row.get("nome"):
        row["nome"] = title.get_text(strip=True).split("|")[0].split("–")[0].strip()

    # Pares label → valor  (dl/dt/dd, tabelas, campos de ficha)
    pares: list[tuple] = []
    for dl in soup.find_all("dl"):
        for dt, dd in zip(dl.find_all("dt"), dl.find_all("dd")):
            t1 = dt.get_text(" ", strip=True)
            t2 = dd.get_text(" ", strip=True)
            if t1 and t2 and t1 != t2:
                pares.append((t1, t2))
    for tr in soup.find_all("tr"):
        for th, td in zip(tr.find_all("th"), tr.find_all("td")):
            t1 = th.get_text(" ", strip=True)
            t2 = td.get_text(" ", strip=True)
            if t1 and t2 and t1 != t2:
                pares.append((t1, t2))
    for el in soup.find_all(class_=re.compile(
            r"field|detail|info|meta|attr|prop|label|ficha|dado", re.I)):
        lbl   = el.find(["label","strong","b","dt","th","span","h4","h5"],
                        class_=re.compile(r"label|key|title|name|campo", re.I))
        val_e = el.find(["span","p","div","dd","td"],
                        class_=re.compile(r"value|content|data|text|body|valor", re.I))
        if lbl and val_e:
            t1 = lbl.get_text(" ", strip=True)
            t2 = val_e.get_text(" ", strip=True)
            if t1 and t2 and t1 != t2:
                pares.append((t1, t2))

    # Mapeamento de sinónimos de labels para colunas fixas
    _label_map = {
        "municipio": ["município","municipio","concelho"],
        "morada":    ["morada","endereço","address","localização"],
        "telefone":  ["telefone","telef","tel ","contacto telefónico"],
        "email":     ["e-mail","email","correio"],
        "horario":   ["horário","horario","horas de funcionamento"],
        "website":   ["website","site","página web"],
        "preco":     ["preço","preço/pessoa","entrada","admissão","bilhete"],
        "latitude":  ["latitude","lat"],
        "longitude": ["longitude","lon","lng"],
    }
    for lbl_raw, val_raw in pares:
        lbl_norm = lbl_raw.lower().strip()
        matched = False
        for field, synonyms in _label_map.items():
            if any(s in lbl_norm for s in synonyms):
                if not row.get(field):
                    row[field] = val_raw
                matched = True
                break
        if not matched:
            key = "campo_" + re.sub(r"[^\w]", "_", lbl_raw.lower().strip("_"))[:40]
            if key not in row:
                row[key] = val_raw

    # Imagens de conteúdo — apenas do domínio douroetamega.pt
    _JUNK = ("googlelogo","sunny.png","weather","spinner","loading",
             "placeholder","logo","icon","favicon","avatar","maps.gstatic",
             "maps.googleapis","gstatic.com","googleapis.com")
    imgs = []
    for img in soup.find_all("img", src=True):
        src = img["src"].strip()
        if not src:
            continue
        full = urljoin(BASE_URL, src)
        # Só imagens do próprio site
        if "douroetamega.pt" not in full:
            continue
        # Excluir ícones/junk
        if any(j in full.lower() for j in _JUNK):
            continue
        if full not in imgs:
            imgs.append(full)
    row["imagens"] = " | ".join(imgs[:15])

    # Descrição longa (fallback corpo da página)
    if not row.get("descricao"):
        for tag in soup.find_all(["p","div"],
                                  class_=re.compile(
                                      r"desc|content|body|text|intro|summary"
                                      r"|about|corpo|conteudo|article", re.I)):
            txt = tag.get_text(" ", strip=True)
            if len(txt) > 80:
                row["descricao"] = txt[:3000]
                break

    # Tags / categorias visíveis
    tags = []
    for el in soup.find_all(class_=re.compile(
            r"\btag\b|\bbadge\b|\bchip\b|\bcategory\b|\bcategoria\b", re.I)):
        t = el.get_text(" ", strip=True)
        if t and len(t) < 80:
            tags.append(t)
    row["tags"] = " | ".join(dict.fromkeys(tags))

    # Redes sociais
    sociais = []
    for a in soup.find_all("a", href=True):
        href = a["href"]
        if any(s in href for s in ["facebook.com","instagram.com","twitter.com",
                                    "youtube.com","tiktok.com","linkedin.com",
                                    "x.com/","threads.net"]):
            if href not in sociais:
                sociais.append(href)
    row["redes_sociais"] = " | ".join(sociais)

    # Município via regex no texto corrido
    raw_text = soup.get_text(" ")
    for pat in [r"munic[íi]pio[:\s]+([A-ZÀ-Úa-zà-ú][^\n<.,]{2,40})",
                r"concelho[:\s]+([A-ZÀ-Úa-zà-ú][^\n<.,]{2,40})"]:
        m = re.search(pat, raw_text, re.I)
        if m and not row.get("municipio"):
            row["municipio"] = m.group(1).strip()

    return row


# ── Excel ─────────────────────────────────────────────────────────────────────

_COLS_PRIORITY = [
    "id", "nome", "h1", "secao", "categoria", "subcategoria", "slug", "tipo",
    "descricao", "municipio", "morada", "localidade", "regiao", "pais",
    "latitude", "longitude", "telefone", "email", "website",
    "preco", "horario",
    "data_inicio", "data_fim", "local_evento",
    "data_publicacao", "data_modificacao", "keywords",
    "imagem_jsonld", "og_image", "og_tipo",
    "imagens", "tags", "redes_sociais",
    "profundidade", "url",
]

_HDR_FILL = PatternFill("solid", fgColor="17375E")
_HDR_FONT = Font(color="FFFFFF", bold=True)
_X_FILL   = PatternFill("solid", fgColor="E2EFDA")


def _write_sheet(ws, headers, rows, presence=False):
    ws.append(headers)
    for i in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=i)
        c.fill = _HDR_FILL
        c.font = _HDR_FONT
        c.alignment = Alignment(horizontal="center")
    for row in rows:
        vals = []
        for col in headers:
            v = str(row.get(col, "") or "").strip()
            vals.append("X" if (presence and v) else ("" if presence else v))
        ws.append(vals)
    if presence:
        for r in ws.iter_rows(min_row=2):
            for cell in r:
                if cell.value == "X":
                    cell.fill = _X_FILL
                    cell.alignment = Alignment(horizontal="center")
    for i, name in enumerate(headers, 1):
        mx = len(name)
        for r in ws.iter_rows(min_row=2, min_col=i, max_col=i):
            for cell in r:
                if cell.value:
                    mx = max(mx, min(len(str(cell.value)), 60))
        ws.column_dimensions[get_column_letter(i)].width = mx + 2


def save_excel(dados, output=OUTPUT):
    if not dados:
        print("[Erro] Sem dados para guardar.")
        return

    all_keys, seen = [], set()
    for row in dados:
        for k in row:
            if k not in seen:
                seen.add(k)
                all_keys.append(k)

    priority = [c for c in _COLS_PRIORITY if c in seen]
    headers  = priority + [c for c in all_keys if c not in priority]

    wb = Workbook()

    ws = wb.active
    ws.title = "Dados"
    _write_sheet(ws, headers, dados)

    ws2 = wb.create_sheet("Presença")
    _write_sheet(ws2, headers, dados, presence=True)

    ws3 = wb.create_sheet("Resumo")
    ws3.append(["secao", "categoria", "total"])
    for i in range(1, 4):
        c = ws3.cell(row=1, column=i)
        c.fill = _HDR_FILL; c.font = _HDR_FONT
        c.alignment = Alignment(horizontal="center")
    counter = Counter(
        (str(r.get("secao", "")), str(r.get("categoria", ""))) for r in dados
    )
    for (s, c), n in sorted(counter.items()):
        ws3.append([s, c, n])
    for i in range(1, 4):
        ws3.column_dimensions[get_column_letter(i)].width = 28

    wb.save(output)
    print(f"\n✅  {output}")
    print(f"    {len(dados)} linhas × {len(headers)} colunas")
    print(f"    Sheets: Dados | Presença (X/vazio) | Resumo")


# ── Download de fotos ─────────────────────────────────────────────────────────

def _foto_session():
    try:
        import cloudscraper
        s = cloudscraper.create_scraper()
    except ImportError:
        import requests
        s = requests.Session()
    s.headers.update({"User-Agent": _session.headers["User-Agent"]})
    return s

def _slug_safe(s: str) -> str:
    return re.sub(r"[^\w-]", "_", s or "geral")[:60].strip("_") or "geral"

def download_fotos(dados: list[dict]):
    """Descarrega todas as imagens para FOTOS_DIR/{secao}/{categoria}/{slug}/"""
    sess   = _foto_session()
    total  = sum(1 for r in dados if r.get("imagens"))
    count  = 0
    errors = 0

    print(f"\n[Fotos] A descarregar imagens de {total} items para {FOTOS_DIR}/")

    for row in dados:
        imgs_raw = row.get("imagens", "")
        if not imgs_raw:
            continue

        secao = _slug_safe(row.get("secao", ""))
        cat   = _slug_safe(row.get("categoria", ""))
        slug  = _slug_safe(row.get("slug", "") or row.get("id", "") or row.get("nome", "item"))
        folder = os.path.join(FOTOS_DIR, secao, cat, slug)
        os.makedirs(folder, exist_ok=True)

        for img_url in imgs_raw.split(" | "):
            img_url = img_url.strip()
            if not img_url:
                continue
            fname = img_url.split("/")[-1].split("?")[0] or "foto.jpg"
            # Garante extensão
            if "." not in fname[-6:]:
                fname += ".jpg"
            fpath = os.path.join(folder, fname)
            if os.path.exists(fpath):
                continue  # já descarregado
            try:
                r = sess.get(img_url, timeout=20)
                if r.ok:
                    with open(fpath, "wb") as f:
                        f.write(r.content)
                    count += 1
            except Exception:
                errors += 1
            time.sleep(DELAY_FOTO)

    print(f"[Fotos] {count} ficheiros descarregados  |  {errors} erros")
    print(f"[Fotos] Pasta: {os.path.abspath(FOTOS_DIR)}/")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("  douroetamega.pt — Crawler v3 (BFS + extracção)")
    print("=" * 60)

    # ── Fase 1: BFS ──────────────────────────────────────────────────────────
    print("\n[Fase 1] BFS — a descobrir todos os URLs de items…\n")
    urls = discover_urls()

    if not urls:
        print("[Aviso] Nenhum URL descoberto.")
        return

    # ── Fase 2: Extracção ────────────────────────────────────────────────────
    total = len(urls)
    print(f"\n[Fase 2] A extrair dados de {total} items…\n")
    dados = []

    for i, url in enumerate(urls, 1):
        if i <= 10 or i % 50 == 0:
            slug = url.rstrip("/").split("/")[-1] or url.rstrip("/").split("/")[-2]
            print(f"  [{i:5}/{total}] {slug}")

        html = _get(url)
        if html:
            row = extract_page(url, html)
            dados.append(row)

        # Guarda parcialmente a cada 300 items
        if i % 300 == 0 and dados:
            save_excel(dados, OUTPUT.replace(".xlsx", f"_parcial_{i}.xlsx"))
            print(f"  [Parcial] {i} items guardados")

        time.sleep(DELAY_EXT)

    save_excel(dados)
    download_fotos(dados)


if __name__ == "__main__":
    main()
